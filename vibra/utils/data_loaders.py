
import h5py
import numpy as np

from openpyxl import load_workbook
from pandas import read_excel
from pathlib import Path


def load_spreadsheet_data(path: str) -> dict:

    imported_results = dict()

    if not Path(path).exists():
        return imported_results

    wb = load_workbook(path)

    skiprows = 0
    sheetnames = wb.sheetnames

    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0, 1, 2]
                                    ).to_numpy()

        except:
            sheet_data = read_excel(
                                    path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0, 1]
                                    ).to_numpy()
            
        filtered_data = [row_data for row_data in sheet_data if not isinstance(row_data[0], str)]
        sheet_data = np.array(filtered_data, dtype=float)

        imported_results[sheetname] = sheet_data

    return imported_results

def load_simulation_data_from_hdf_file(path: str | Path):

    simulation_data = dict()

    with h5py.File(path, 'r') as hf:

        # Read key data
        nodal_data = hf.get("nodal_data")
        variables = hf.get("variables")
        metadata = hf.get("metadata")

        # Save the nodal coordinates matrix
        simulation_data["nodal_area"] = np.array(nodal_data.get("nodal_area"), dtype=float)
        simulation_data["nodal_coordinates"] = np.array(nodal_data.get("coords"), dtype=float)

        # Save other attributes
        for attr_key in metadata.attrs:
            simulation_data[attr_key] = metadata.attrs[attr_key]

        # Calculate the start point from last revolution
        delta_theta = metadata.attrs["delta_theta"]
        steps_per_rev = int(360 / delta_theta)
        start = steps_per_rev + 1

        # filter the last revolution data for metadata
        for key, values in metadata.items():
            simulation_data[key] = values[-start:]

        # filter the last revolution data for variables
        for key, values in variables.items():
            simulation_data[key] = values[:, -start:]

    return simulation_data