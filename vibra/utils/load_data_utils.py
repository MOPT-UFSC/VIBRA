
from pandas import read_excel
from openpyxl import load_workbook

import numpy as np
from pathlib import Path


def load_spreadsheet_data(path: str) -> dict:

    imported_results = dict()

    if not Path(path).exists():
        return imported_results

    wb = load_workbook(path)

    skiprows = 0
    sheetnames = wb.sheetnames

    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0, 1, 2]
                                    ).to_numpy()

        except:
            sheet_data = read_excel(
                                    path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0, 1]
                                    ).to_numpy()
            
        filtered_data = [row_data for row_data in sheet_data if not isinstance(row_data[0], str)]
        sheet_data = np.array(filtered_data, dtype=float)

        imported_results[sheetname] = sheet_data

    return imported_results