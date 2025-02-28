from PySide6.QtWidgets import QComboBox, QDialog, QFileDialog, QLabel, QLineEdit, QPushButton
from PySide6.QtGui import QCloseEvent, QIcon
from PySide6.QtCore import Qt
from pathlib import Path

from vibra import app, UI_DIR
from vibra.interface.formatters.icons import *
from vibra.interface.general.print_message_input import PrintMessageInput
from vibra.interface.general.get_user_confirmation_input import GetUserConfirmationInput
from vibra.interface.formatters.config_widget_appearance import ConfigWidgetAppearance
from molde import load_ui

import os


class LoadFluidCompositionInput(QDialog):
    def __init__(self, *args, **kwargs):
        super().__init__()

        ui_path = UI_DIR / "model/setup/fluid/load_fluid_composition.ui"
        load_ui(ui_path, self, UI_DIR)

        self.main_window = app().main_window
        self.main_window.set_input_widget(self)

        self.file_path = kwargs.get("file_path", "")
       
        self._initialize()
        self._config_window()
        self._define_qt_variables()
        self._create_connections()

        ConfigWidgetAppearance(self, tool_tip=True)

        self._config_widgets()
        self._load_file()
        self.exec()

    def _initialize(self):

        self.complete = False
        self.fluid_composition_data = None

        user_path = os.path.expanduser('~')
        desktop_path = Path(os.path.join(os.path.join(user_path, 'Desktop')))
        self.desktop_path = str(desktop_path)
        
    def _config_window(self):
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowIcon(app().main_window.vibra_icon)
        self.setWindowTitle(f"Vibra")

    def _define_qt_variables(self):

        # QComboBox
        self.comboBox_sheet_names : QComboBox

        # QLineEdit
        self.lineEdit_file_path : QLineEdit

        # QPushButton
        self.pushButton_exit : QPushButton
        self.pushButton_confirm : QPushButton
        self.pushButton_search : QPushButton

    def _create_connections(self):
        self.pushButton_exit.clicked.connect(self.close)
        self.pushButton_confirm.clicked.connect(self.confirm_button_callback)
        self.pushButton_search.clicked.connect(self.search_button_callback)

    def _config_widgets(self):
        self.lineEdit_file_path.setDisabled(True)
        self.comboBox_sheet_names.setDisabled(True)

    def _load_file(self):
        if os.path.exists(self.file_path):
            self.lineEdit_file_path.setText(self.file_path)
            self.load_composition_data_from_file()

    def search_button_callback(self):

        last_geometry_file = app().config.get_last_folder_for("fluid_composition_folder")
        if last_geometry_file is None:
            inital_path = self.desktop_path
        else:
            inital_path = last_geometry_file

        file_path, check = QFileDialog.getOpenFileName( None,
                                                        'Open file',
                                                        inital_path,
                                                        'Files (*.xlsx *.xls)')

        if not check:
            return True
        
        self.file_path = file_path

        app().config.write_last_folder_path_in_file("fluid_composition_folder", file_path)
        self.lineEdit_file_path.setText(file_path)

        self.load_composition_data_from_file()

    def load_composition_data_from_file(self):

        if self.lineEdit_file_path.text() == "":
            return

        self.imported_data = dict()
        self.comboBox_sheet_names.clear()

        from pandas import read_excel
        from openpyxl import load_workbook

        wb = load_workbook(self.file_path)
        sheetnames = wb.sheetnames
        for sheetname in sheetnames:

            try:
                sheet_data = read_excel(
                                        self.file_path, 
                                        sheet_name = sheetname, 
                                        header = 0, 
                                        usecols = [0,1,2,3]
                                        ).to_numpy()
                
                self.imported_data[sheetname] = sheet_data
                self.comboBox_sheet_names.addItem(sheetname)
                
            except Exception as error_log:
                window_title = "Error"
                title = "Error while reading data from file"
                message = f"{str(error_log)}"
                PrintMessageInput([window_title, title, message])
                return True

        self.comboBox_sheet_names.setDisabled(False)
               
    def confirm_button_callback(self):
        if self.imported_data:
            selection = self.comboBox_sheet_names.currentText()
            self.fluid_composition_data = self.imported_data[selection]
            self.complete = True
            self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.confirm_button_callback()
            return
        elif event.key() == Qt.Key_Escape:
            self.close()