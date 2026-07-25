import os
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QFileDialog

from vibra import app
from vibra.interface.general.print_message_input import PrintMessageInput
from vibra.interface.ui_generated.model.fluid.load_fluid_composition_ui import LoadFluidComposition_UI


class LoadFluidCompositionInputs(LoadFluidComposition_UI):
    def __init__(self, file_path: str = ""):
        super().__init__()

        app().main_window.set_input_widget(self)

        self.file_path = file_path
       
        self._initialize()
        self._config_window()
        self._create_connections()

        self._config_widgets()
        self._load_file()
        self.exec()

    def _initialize(self):

        self.complete = False
        self.fluid_composition_data: list[tuple[int, str, str, str]] = []

        desktop_path = Path.home() / "Desktop"
        self.desktop_path = str(desktop_path)
        
    def _config_window(self):
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowIcon(app().main_window.vibra_icon)
        self.setWindowTitle("Vibra")

    def _create_connections(self):
        self.pushButton_cancel.clicked.connect(self.close)
        self.pushButton_load_composition.clicked.connect(self.confirm_button_callback)
        self.pushButton_search.clicked.connect(self.search_button_callback)

    def _config_widgets(self):
        self.lineEdit_file_path.setDisabled(True)
        self.comboBox_sheet_names.setDisabled(True)

    def _load_file(self):
        if os.path.exists(self.file_path):
            self.lineEdit_file_path.setText(self.file_path)
            self.load_composition_data_from_file()

    def search_button_callback(self):

        last_geometry_file = app().config.get_last_folder_for("fluid_composition_folder")
        if last_geometry_file is None:
            initial_path = self.desktop_path
        else:
            initial_path = last_geometry_file

        file_path, check = QFileDialog.getOpenFileName(
            None,
            "Open file",
            str(initial_path),
            "Files (*.xlsx *.xls)",
        )

        if not check:
            return True
        
        self.file_path = file_path

        app().config.write_last_folder_path_in_file("fluid_composition_folder", file_path)
        self.lineEdit_file_path.setText(file_path)

        self.load_composition_data_from_file()

    def load_composition_data_from_file(self):

        if self.lineEdit_file_path.text() == "":
            return

        self.imported_data = {}
        self.comboBox_sheet_names.clear()

        from openpyxl import load_workbook
        from polars import read_excel

        wb = load_workbook(self.file_path)
        sheetnames = wb.sheetnames

        for sheetname in sheetnames:

            try:
                sheet_data = read_excel(
                    self.file_path,
                    sheet_name=sheetname,
                    columns=(0, 1, 2, 3),
                    has_header=True,
                )

                self.imported_data[sheetname] = sheet_data.to_numpy()
                self.comboBox_sheet_names.addItem(sheetname)
               
            except Exception as error_log:
                window_title = "Error"
                title = "Error while reading data from file"
                message = str(error_log)
                PrintMessageInput([window_title, title, message])
                return True

        self.comboBox_sheet_names.setDisabled(False)
               
    def confirm_button_callback(self):
        if self.imported_data:
            selection = self.comboBox_sheet_names.currentText()
            self.fluid_composition_data = self.imported_data[selection]
            self.complete = True
            self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.confirm_button_callback()
            return
        elif event.key() == Qt.Key_Escape:
            self.close()