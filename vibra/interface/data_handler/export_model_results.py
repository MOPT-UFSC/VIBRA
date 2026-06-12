import logging
import platform
from pathlib import Path

import numpy as np
from PySide6.QtWidgets import QFileDialog

from vibra import app


class ExportModelResults(QFileDialog):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self._initialize()

    def _initialize(self):
        self.data = dict()

    def _set_data_to_export(self, data: dict, **kwargs):
        self.data = data
        if data:
            self.call_file_dialog_and_export_data(**kwargs)

    def export_data_in_text_format(self, export_path: str, delimiter: str = ","):

        for data in self.data.values():

            if not isinstance(data, dict):
                continue

            unit = data["unit"]
            x_data = data["x_data"]
            y_data = data["y_data"]
            x_label = data.get("x_label")
            y_label = data.get("y_label")

            if isinstance(y_data[0], complex):
                header = [x_label, f"{y_label} - real [{unit}]", f"{y_label} - imaginary [{unit}]", f"{y_label} - absolute [{unit}]"]
                header = f"{x_label}, {y_label} - real [{unit}], {y_label} - imaginary [{unit}], Absolute [{unit}]"
                data_to_export = np.array([x_data, np.real(y_data), np.imag(y_data), np.abs(y_data)]).T

            else:
                data_type = data.get("data_type", "")
                header = f"{x_label}, {data_type.capitalize()} [{unit}]"
                data_to_export = np.array([x_data, y_data]).T

            logging.info("Exporting data... (85%)")
            np.savetxt(export_path, data_to_export, delimiter=delimiter, header=header)

    def export_data_in_spreadsheet_format(self, export_path: str, **kwargs):

        from openpyxl import load_workbook
        from pandas import ExcelWriter
        from polars import DataFrame, read_excel

        logging.info("Exporting data... (75%)")

        existing_data_frames = dict()
        existing_path = kwargs.get("existing_path", "")

        if Path(existing_path).exists():
            ext = existing_path.split(".")[-1]
            if ext in ["xls", "xlsx"]:
                wb = load_workbook(existing_path)
                sheetnames = wb.sheetnames

                for sheet_name in sheetnames:
                    existing_data_frames[sheet_name] = read_excel(existing_path, sheet_name=sheet_name, columns=[0, 1, 2, 3], engine="openpyxl")

        logging.info("Exporting data... (85%)")

        with ExcelWriter(export_path) as writer:

            for key, existing_df in existing_data_frames.items():
                existing_df: DataFrame
                existing_df.to_pandas().to_excel(writer, sheet_name=key, index=False)

            count = 0
            for key, data in self.data.items():

                if not isinstance(data, dict):
                    continue

                if len(key) == 2:
                    if key[1] is None:
                        sheet_name = f"{key[0]}"
                    else:
                        selection_type, selection_id = key
                        sheet_name = f"{selection_type}_{selection_id}"
                else:
                    count += 1
                    sheet_name = f"sheet_{count}"

                unit = data.get("unit")
                x_data = data.get("x_data")
                y_data = data.get("y_data")
                x_label = data.get("x_label")
                y_label = data.get("y_label")

                if isinstance(y_data[0], complex):
                    header = [x_label, f"{y_label} - real [{unit}]", f"{y_label} - imaginary [{unit}]", f"Absolute [{unit}]"]
                    data_to_export = np.array([x_data, np.real(y_data), np.imag(y_data), np.abs(y_data)]).T
 
                else:
                    data_type = data["data_type"]
                    header = [x_label, f"{data_type.capitalize()} [{unit}]"]
                    data_to_export = np.array([x_data, y_data]).T

                df = DataFrame(data_to_export, schema=header)
                df.to_pandas().to_excel(writer, sheet_name=sheet_name, index=False)

    def call_file_dialog_and_export_data(self, **kwargs):

        existing_path = kwargs.get("existing_path", "")

        if existing_path:
            file_path = existing_path

        else:
            caption = "Export the model results"

            path = app().config.get_last_folder_for("exported_data_folder")
            if path is None:
                directory_path = Path().home()
            else:
                directory_path = path

            if len(self.data) == 1:
                _filter = "Spreadsheet (*.xlsx);; Spreadsheet (*.xls);; Text file (*.dat);; Text file (*.txt);; Text file (*.csv)"
            else:
                _filter = "Spreadsheet (*.xlsx)"

            kwargs = dict()
            if platform.system() == "Linux":
                kwargs["options"] = QFileDialog.Option.DontUseNativeDialog

            file_path, selected_filter = self.getSaveFileName(app().main_window, caption, str(directory_path), filter=_filter,**kwargs)
            if not file_path:
                return

            file_path = Path(file_path)

            if not file_path.suffix:
                file_extension = f".{self.get_file_extension(selected_filter)}"
                file_path = file_path.with_suffix(file_extension)

        app().config.write_last_folder_path_in_file("exported_data_folder", file_path)

        logging.info("Exporting data... (50%)")
        if file_path.suffix.lower() in [".xls", ".xlsx"]:
            self.export_data_in_spreadsheet_format(str(file_path), existing_path=existing_path)
        else:
            self.export_data_in_text_format(str(file_path))

    def get_file_extension(self, check: str) -> str:
        return check.split(".")[1][:-1]