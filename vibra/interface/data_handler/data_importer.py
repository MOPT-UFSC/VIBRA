from PySide6.QtWidgets import QFileDialog

from vibra import app
from vibra.interface.data_handler.imported_data import ImportedData

from typing import List
from pathlib import Path
import numpy as np
import platform
import os


class DataImporter:

    @staticmethod
    def __import_files(caption: str, last_folder: str, file_extensions: List[str], multiple_files: bool = False):

        folder_path = app().config.get_last_folder_for(last_folder)
        if folder_path is None:
            folder_path = os.path.expanduser("~")

        kwargs = dict()
        if platform.system() == "Linux":
                kwargs["options"] = QFileDialog.Option.DontUseNativeDialog

        str_extensions = "Files ("
        for extension in file_extensions:
            str_extensions += "*."
            str_extensions += extension
            str_extensions += " "
        
        str_extensions = str_extensions.strip()
        str_extensions += ")"

        imported_paths, file_extension = None, None
        if (multiple_files):
            imported_paths, file_extension = QFileDialog.getOpenFileNames( None, 
                                                                caption, 
                                                                folder_path, 
                                                                str_extensions,
                                                                **kwargs)
        else:
            imported_paths, file_extension = QFileDialog.getOpenFileName( None, 
                                                                caption, 
                                                                folder_path, 
                                                                str_extensions,
                                                                **kwargs)
            
        if "h5" in file_extensions or "hdf5" in file_extensions:
            return imported_paths, file_extension

        if not file_extension:
            return
        
        imported_data = None
        last_imported_file = imported_paths if isinstance(imported_paths, str) else imported_paths[-1]

        imported_data = list()
        if isinstance(imported_paths, list):
            for imported_path in imported_paths:
                imported_data.extend(DataImporter.read_data_in_file(imported_path, use_first_sheet=False))

        else:
            imported_data.extend(DataImporter.read_data_in_file(imported_paths, use_first_sheet=True))

        app().config.write_last_folder_path_in_file(last_folder, last_imported_file)
        
        return imported_data

    # @staticmethod
    # def get_file_paths(caption: str, last_folder: str, file_extensions: List[str], multiple_files: bool = False):

    #     folder_path = app().config.get_last_folder_for(last_folder)
    #     if folder_path is None:
    #         folder_path = os.path.expanduser("~")

    #     kwargs = dict()
    #     if platform.system() == "Linux":
    #             kwargs["options"] = QFileDialog.Option.DontUseNativeDialog

    #     str_extensions = "Files ("
    #     for extension in file_extensions:
    #         str_extensions += "*."
    #         str_extensions += extension
    #         str_extensions += " "
        
    #     str_extensions = str_extensions.strip()
    #     str_extensions += ")"

    #     if (multiple_files):
    #         imported_paths, file_extension = QFileDialog.getOpenFileNames( None, 
    #                                                             caption, 
    #                                                             folder_path, 
    #                                                             str_extensions,
    #                                                             **kwargs)
    #     else:
    #         imported_paths, file_extension = QFileDialog.getOpenFileName( None, 
    #                                                             caption, 
    #                                                             folder_path, 
    #                                                             str_extensions,
    #                                                             **kwargs)
            
    #     if not file_extension:
    #         return None, None

    #     return imported_paths, file_extension
  
    @staticmethod
    def import_multiple_files(last_folder: str, file_extensions: List[str], caption: str = "Open file") -> List[ImportedData]:
        return DataImporter.__import_files(caption, last_folder, file_extensions, True)

    @staticmethod
    def import_single_file(last_folder: str, file_extensions: List[str], caption: str = "Open File") -> ImportedData | None:
        imported_data = DataImporter.__import_files(caption, last_folder, file_extensions)
        if "hdf" in file_extensions or "hdf5" in file_extensions:
            return imported_data

        if isinstance(imported_data, list):
            if imported_data:
                return imported_data[0]
        return None

    @staticmethod
    def read_data_in_file(file_path: str, use_first_sheet: bool = True):

        import warnings

        output_data = list()

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore")
               
            sufix = Path(file_path).suffix
            filename = os.path.basename(file_path)

            if sufix in [".txt", ".dat", ".csv"]:
                try:
                    loaded_data = np.loadtxt(file_path, delimiter = ",")
                except:
                    loaded_data = DataImporter.__load_text_file_data(file_path)

                loaded_data = DataImporter.__remove_unnecesary_header_in_data(loaded_data)
                output_data.append(ImportedData(loaded_data, filename, sufix, path=file_path))
                
            elif sufix in [".xls", ".xlsx"]:

                from pandas import read_excel
                from openpyxl import load_workbook

                wb = load_workbook(file_path)
                
                for sheetname in wb.sheetnames:
                    for cols in [(0, 1, 2), (0, 1)]:
                        try:
                            sheet_data = read_excel(
                                                    file_path, 
                                                    sheet_name = sheetname,  
                                                    usecols = cols,
                                                    engine = "openpyxl",
                                                    ).to_numpy()
                            break
                        except:
                            pass

                    sheet_data = DataImporter.__remove_unnecesary_header_in_data(sheet_data)
                    output_data.append(ImportedData(sheet_data, filename, sufix, sheetname, file_path))
                    if use_first_sheet:
                        break

            return output_data

    @staticmethod                      
    def __remove_unnecesary_header_in_data(data: np.ndarray) -> np.ndarray:
        filtered_data = [row for row in data if not isinstance(row[0], str)]
        return np.array(filtered_data, dtype=float)

    @staticmethod
    def __load_text_file_data(path: str):

        output_data = list()
        if isinstance(path, str):
            path = Path(path)

        with open(path, 'r') as file:
            for line in file.readlines():
                try:
                    modif_line = line.replace(",", " ").strip().split(" ")
                    line_values = [float(value) for value in modif_line if value != ""]
                except:
                    continue

                output_data.append(line_values)

        return np.array(output_data, dtype=float)