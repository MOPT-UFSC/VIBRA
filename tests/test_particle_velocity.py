from vibra.engine.properties.fluid import Fluid
from vibra.engine.mesher.mesh import Mesh
from vibra.engine.mesher.element_type import *
from vibra.engine.model import Model
from vibra.engine.assemblers.acoustic_assembler import AcousticAssembler
from vibra.engine.solvers.acoustic_modal_solver import AcousticModalSolver
from vibra.engine.solvers.acoustic_harmonic_solver import AcousticHarmonicSolver
from vibra.external_mesh.external_mesh_data import ExternalMeshData

import numpy as np
import matplotlib.pyplot as plt
from time import time

import os
import openpyxl
import pandas as pd

def test_load_external_mesh_and_solve(reorder_nodes=False):
    return

    mesh_path = "validation/data/particle_velocity/mesh/ds_tubo_reto.dat"

    if not os.path.exists(mesh_path):
        return

    # define the known 'Named selections' from model
    named_selecion_to_tag = { 
                                "inlet_face" : 1,
                                "outlet_face" : 2
                            }

    t0 = time()
    external_mesh = ExternalMeshData()
    external_mesh.reset()
    external_mesh.read_file(mesh_path)
    external_mesh.set_named_selections(list(named_selecion_to_tag.keys()))
    external_mesh.decode_mesh_data_from_file()

    mesh = Mesh()
    mesh.import_external_nodal_coordinates(external_mesh.nodal_coordinates, index_zero=True)
    mesh.import_external_connectivity(external_mesh.connectivity_arrays, index_zero=True, etype_tag=4)
    # mesh.export_nodal_coordinates("nodal_coordinates.dat")
    # mesh.export_solid_elements_connectivity("solids_connectivity.dat")
    mesh.element_type = TETRAHEDRON_4

    for named_selection, surf_data in external_mesh.elements_from_named_selection.items():
        tag = named_selecion_to_tag[named_selection]
        mesh.elements_from_surface[tag] = surf_data["element_indexes"]
        mesh.connectivity_from_surfaces[tag] = surf_data["connectivity"] - 1
        ns_nodes = external_mesh.nodes_from_named_selection[named_selection]
        mesh.nodes_from_surfaces[tag] = np.array(ns_nodes, dtype=int) - 1

        mesh.volume_from_surface[tag] = [1]

    mesh.surfaces_from_volumes[1] = [1, 2]

    # if reorder_nodes:
    #     mesh._process_nodes_reordering()
    #     map_nodes_indexes = mesh.reordering.map_nodes_indexes

    # Define the fluid properties
    rho_0 = 1.18
    c_0 = 343.0
    mu = 0*1.8e-05
    #
    fluid = Fluid(  name = "Air",
                    identifier = 1,
                    color = (200, 200, 200),
                    fluid_density = rho_0,
                    speed_of_sound = c_0,
                    dynamic_viscosity = mu  )

    # Set the defined fluid
    model = Model()
    model.mesh =  mesh
    model.generated_mesh = True
    model.set_fluid(fluid, volume=1)

    # Normal surface velocity data
    data_Vn = { "real_values" : [1],
                "imag_values" : [0],
                "nodal_attribution" : False,
                "averaged" : False }

    # Impedance data
    Zo = fluid.impedance
    data_Z = {  "real_values" : [Zo],
                "imag_values" : [0],
                "nodal_attribution" : False,
                "averaged" : False  }

    model.set_surface_velocity(data_Vn, 2)
    model.set_specific_impedance(data_Z, 1)
    model.set_specific_impedance(data_Z, 2)

    # Create the acoustic assembler
    assembler = AcousticAssembler(model)

    # Define the analysis frequency setup
    df = 1
    f_min = 1
    f_max = 500
    frequencies = np.arange(f_min, f_max + df, df)

    # Set the analysis frequency setup
    assembler.set_frequencies(frequencies)
    assembler.process_assemble()
    
    # t0 = time()
    # # Run modal analysis
    # modal_solver = AcousticModalSolver(assembler)
    # natural_frequencies, modal_shape = modal_solver.solve()
    # dt = time() - t0
    # print(f"Elapsed time to solve modal analysis: {round(dt, 4)}s")
    # return
    
    # Define the analysis type and load setup
    analysis_data = {"analysis_id" : 3, "frequencies" : frequencies}
    harmonic_solver = AcousticHarmonicSolver(assembler, analysis_data=analysis_data)
    
    t0 = time()
    solution = harmonic_solver.solve(print_log=True)
    dt = time() - t0
    print(f"Elapsed time to solve harmonic analysis: {round(dt, 4)}")

    element_id = 5648 - 1
    node_id = 13 - 1

    element_id = 1750 - 1
    node_id = 1198 - 1

    element_3d, _ = assembler.get_element()
    element_3d.reorder_connect()

    # Vxyz = element_3d.process_particle_velocity(element_id, node_id, rho_0, frequencies, solution)
    # print(Vxyz[0, :])

    elements_connected_to_nodes =  mesh.get_solid_elements_connected_to_nodes([node_id])
    # print(elements_connected_to_nodes)

    data = dict()
    for _node_id, element_ids in elements_connected_to_nodes.items():
        Vk = 0.
        for _element_id in element_ids:
            Vk += element_3d.process_particle_velocity(_element_id, _node_id, rho_0, frequencies, solution)
        data[_node_id] = Vk / len(element_ids)

    Vx = data[node_id][0, :]
    # print(data[_node_id][0, :])

    dt = time() - t0
    print(f"Elapsed time to solve harmonic analysis: {round(dt, 4)}")

    if solution is not None:

        # cols = solution.shape[1]       
        # results = np.zeros((cols, 3), dtype=float)
        # results[:, 0] = frequencies
        # results[:, 1] = np.real(solution[node, :])
        # results[:, 2] = np.imag(solution[node, :])
        # filename = f"acoustic_pressure_at_node_{node}_Vibra_pardiso.dat"
        # np.savetxt(filename, results, delimiter=",")

        external_results = get_external_results()

        data = external_results[f"pressure_node_{node_id+1}"]

        freq_ref = data["x_data"]
        P_ref = data["y_data"]

        # error_abs = np.abs((solution[node, :] - P_ref)/((solution[node, :] + P_ref)/2))
        # assert error_abs < 1e-1

        fig1, ax1 = plt.subplots()
        ax1.semilogy(frequencies, np.abs(solution[node_id, :]), 'r', label='VIBRA')
        ax1.semilogy(freq_ref, np.abs(P_ref), 'k--', label='ANSYS')
        # ax1.semilogy(freq_ref, np.abs(solution[node_id, :] - P_ref), 'k--', label='ANSYS')
        # ax1.semilogy(freq_ref, error_abs, 'k--', label='ANSYS')
        ax1.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Absolute', title='Harmonic Response - Outlet pressure')
        ax1.grid()

        fig2, ax2 = plt.subplots()
        ax2.plot(frequencies, np.real(solution[node_id, :]), 'r', label='VIBRA')
        ax2.plot(freq_ref, np.real(P_ref), 'k--', label='ANSYS')
        ax2.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title='Harmonic Response - Outlet pressure')
        ax2.grid()

        fig3, ax3 = plt.subplots()
        ax3.plot(frequencies, np.imag(solution[node_id, :]), 'r', label='VIBRA')
        ax3.plot(freq_ref, np.imag(P_ref), 'k--', label='ANSYS')
        ax3.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title='Harmonic Response - Outlet pressure')
        ax3.grid()

        data = external_results[f"velocity_node_{node_id+1}"]

        freq_ref = data["x_data"]
        Vx_ref = data["y_data"]

        fig4, ax4 = plt.subplots()
        ax4.plot(frequencies, np.real(Vx), 'r', label='VIBRA')
        ax4.plot(freq_ref, np.real(Vx_ref), 'k--', label='ANSYS')
        ax4.set(xlabel='Frequency [Hz]', ylabel='Particle velocity [m/s] - real', title='Harmonic Response - Inlet Vx')
        ax4.grid()

        plt.legend()
        plt.show()

def get_faces_connectivities():

    connect_face1 = np.array([[1,15,16,1200],    #tubo X
                              [2,16,17,1200],
                              [3,17,1197,1200],
                              [4,17,18,1197],
                              [5,18,19,1197],
                              [6,19,1197,1199],
                              [7,19,20,1199],
                              [8,11,20,1199],
                              [9,11,12,1199],
                              [10,12,1198,1199],
                              [11,12,13,1198],
                              [12,13,14,1198],
                              [13,14,1198,1200],
                              [14,14,15,1200],
                              [15,1197,1198,1200],
                              [16,1197,1198,1199]], dtype=int) - 1
    
    connect_face2 = np.array([  [1,6,7,1203],   #tubo
                                [2,7,8,1203],
                                [3,8,1201,1203],
                                [4,8,9,1201],
                                [5,9,10,1201],
                                [6,10,1201,1204],
                                [7,1,10,1204],
                                [8,1,2,1204],
                                [9,2,3,1204],
                                [10,3,1202,1204],
                                [11,3,4,1202],
                                [12,4,5,1202],
                                [13,5,1202,1203],
                                [14,5,6,1203],
                                [15,1201,1202,1203],
                                [16,1201,1202,1204]  ], dtype=int) - 1

    connectivity_from_surfaces = dict()

    connectivity_from_surfaces[1] = {   "element_indexes" : connect_face1[:, 0],
                                        "connectivity" : connect_face1[:, 1:]   }

    connectivity_from_surfaces[2] = {   "element_indexes" : connect_face2[:, 0],
                                        "connectivity" : connect_face2[:, 1:]   }

    return connectivity_from_surfaces


def get_external_results():
    imported_results = dict()

    # results_path = "validation/data/particle_velocity/results/external_results_Vn1_Z2.xlsx"
    # results_path = "validation/data/particle_velocity/results/external_results_Vn1_Z1_Z2.xlsx"
    # results_path = "validation/data/particle_velocity/results/external_results_Vn2_Z1.xlsx"
    results_path = "validation/data/particle_velocity/results/external_results_Vn2_Z1_Z2.xlsx"

    wb = openpyxl.load_workbook(results_path)

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        sheet_data = pd.read_excel( results_path, 
                                    sheet_name = sheetname, 
                                    header = 0, 
                                    usecols = [0,1,2] ).to_numpy()

        imported_results[sheetname] = {"x_data" : sheet_data[:, 0],
                                       "y_data" : sheet_data[:, 1] + 1j*sheet_data[:, 2]}
        
    return imported_results

def get_solid_elements_connected_to_nodes(solids_connectivity : np.ndarray, node_ids = None, face_connectivity = None):

    solid_elements_connected_to_nodes = dict()

    if isinstance(solids_connectivity, np.ndarray):
        solids_connectivity -= 1

    if isinstance(node_ids, (np.ndarray, list)):
        if isinstance(node_ids, list):
            node_ids = np.array(node_ids, dtype=int)
        selected_ids = node_ids
        selected_ids -= 1

    elif isinstance(face_connectivity, np.ndarray):
        # esta função reordena os elementos do array em um vetor removendo os índices repetidos
        flat_data = (face_connectivity[:, 1:]).flatten()
        selected_ids = np.array([*set(flat_data)], dtype=int)
        selected_ids -= 1

    else:
        print("Insert the list of 'node_ids' or the array with 'face_connectivity' to process solid elements connected to nodes")
        return

    for i, node_id in enumerate(selected_ids):
        # t0 = time()
        mask = np.sum(solids_connectivity[:, 1:] == node_id, axis=1) == 1
        solid_elements_connected_to_nodes[node_id] = solids_connectivity[:, 0][mask]
        # dt = time() - t0
        # print(f"Loop time: {dt} s")

    return solid_elements_connected_to_nodes

def get_particle_velocity(elem_id):
    # aqui deveríamos chamar a função que calcula a Vn em cada elemento
    # seria interessante que ela retorne um array na forma [N_nodes x N_freq]
    return 0

def process_particle_velocity(solid_elements_connected_to_nodes : dict):

    particle_velocity = dict()
    for node_id, elements_from_node in solid_elements_connected_to_nodes.items():

        aux = 0.
        for element_id in elements_from_node:
            aux += get_particle_velocity(element_id)
        
        particle_velocity[node_id] = aux

def plot_results():

    path_pardiso = "temp/acoustic_pressure_at_node_3596_Vibra_pardiso.dat"
    path_scipy = "temp/acoustic_pressure_at_node_3596_Vibra_scipy.dat"

    data_pardiso = np.loadtxt(path_pardiso, delimiter=",")
    data_scipy = np.loadtxt(path_scipy, delimiter=",")

    freq_pardiso = data_pardiso[:, 0]
    Xf_pardiso = data_pardiso[:, 1] + 1j*data_pardiso[:, 2]

    freq_scipy = data_scipy[:, 0]
    Xf_scipy = data_scipy[:, 1] + 1j*data_scipy[:, 2]

    fig, ax1 = plt.subplots()
    ax1.semilogy(freq_pardiso, np.abs(Xf_pardiso), 'r', label='pardiso')
    ax1.semilogy(freq_scipy, np.abs(Xf_scipy), 'k--', label='scipy')
    # ax1.semilogy(freq_scipy, np.abs(Xf_pardiso - Xf_scipy), 'k-', label='difference')
    ax1.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Absolute', title='Harmonic Response - Outlet pressure')
    ax1.grid()
    plt.legend()
    plt.show()

def save_results(data):
    pass

if __name__ == "__main__":
    test_load_external_mesh_and_solve(reorder_nodes=False)
    # plot_results()