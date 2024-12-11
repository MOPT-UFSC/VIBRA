from vibra.engine.properties.fluid import Fluid
from vibra.engine.mesher.mesh import Mesh
from vibra.engine.mesher.element_type import *
from vibra.engine.model import Model
from vibra.engine.assemblers.acoustic_assembler import AcousticAssembler
from vibra.engine.solvers.acoustic_modal_solver import AcousticModalSolver
from vibra.engine.solvers.acoustic_harmonic_solver import AcousticHarmonicSolver
from vibra.external_mesh.external_mesh_data import ExternalMeshData

import os
import pytest

import matplotlib.pyplot as plt
import numpy as np

from pandas import read_excel
from openpyxl import load_workbook

from time import time

pm_model = "DB"

@pytest.mark.slow
def test_load_external_mesh_and_solve():
    return

    # start decoding the Ansys script file (ds.dat file or input file)
    mesh_path = "validation/data/particle_velocity/mesh/silencer/ds_only_fluid_of_silencer_suction_stg1.dat"

    
    if pm_model not in ["DB", "DBM", "JCA"]:
        return

    if not os.path.exists(mesh_path):
        return

    # define the known 'Named selections' from model
    named_selecion_to_tag = { 
                                "input_face" : 1,
                                "output_face" : 2
                            }

    t0 = time()
    external_mesh = ExternalMeshData()
    external_mesh.reset()
    external_mesh.read_file(mesh_path)
    external_mesh.set_named_selections(list(named_selecion_to_tag.keys()))
    external_mesh.decode_mesh_data_from_file()
    
    dt = time() - t0
    print(f"\n\nElapsed time to decode the external mesh data: {round(dt, 4)} s")

    mesh = Mesh()
    mesh.import_external_nodal_coordinates(external_mesh.nodal_coordinates, index_zero=True)
    mesh.import_external_connectivity(external_mesh.connectivity_arrays, index_zero=True, etype_tag=4)
    mesh.export_nodal_coordinates("nodal_coordinates.dat")
    mesh.export_solid_elements_connectivity("solids_connectivity.dat")
    mesh.element_type = TETRAHEDRON_4

    for named_selection, surf_data in external_mesh.elements_from_named_selection.items():
        tag = named_selecion_to_tag[named_selection]
        mesh.elements_from_surface[tag] = surf_data["element_indexes"] - 1
        mesh.connectivity_from_surfaces[tag] = surf_data["connectivity"] - 1
        mesh.nodes_out_of_face_element[tag] = surf_data["outer_nodes"] - 1
        ns_nodes = external_mesh.nodes_from_named_selection[named_selection]
        mesh.nodes_from_surfaces[tag] = np.array(ns_nodes, dtype=int) - 1

        mesh.volume_from_surface[tag] = [1]

    mesh.surfaces_from_volumes[1] = [1, 2]

    # if reorder_nodes:
    #     mesh._process_nodes_reordering()
    #     map_nodes_indexes = mesh.reordering.map_nodes_indexes

    # Define the fluid properties

    temperature = 298.15
    pressure = 122525
    
    rho_0 = 2.634167
    c_0 = 225.307464
    mu = 8.156509e-06
    Cp = 1664.133942
    kt = 1.741553e-02
    gamma = 1.120295
    molar_mass = 51.951533

    fluid = Fluid(  name = "Silencer suction stg1",
                    identifier = 1,
                    color = (200, 200, 200),
                    pressure = pressure,
                    temperature = temperature,
                    fluid_density = rho_0,
                    speed_of_sound = c_0,
                    isentropic_exponent = gamma,
                    thermal_conductivity = kt,
                    specific_heat_Cp = Cp,
                    dynamic_viscosity = mu,
                    molar_mass = molar_mass  )

    # Set the defined fluid
    model = Model()
    model.mesh =  mesh
    model.generated_mesh = True

    for vol_id in [1]:
        model.set_fluid(fluid, volume=vol_id)

    model.set_fluid(fluid, surface=1)
    model.set_fluid(fluid, surface=2)

    # Normal surface velocity data
    data_Vn = { "real_values" : [1],
                "imag_values" : [0],
                "nodal_attribution" : False,
                "averaged" : False }
    
    complex_fluid_data = get_complex_impedance_data()
    impedance_data = complex_fluid_data["complex_impedance"]

    # Impedance data
    # Zo = fluid.impedance
    # data_Z = {  "real_values" : [Zo],
    #             "imag_values" : [0],
    #             "nodal_attribution" : False,
    #             "averaged" : False  }

    data_Z = {  "real_values" : list(impedance_data[:, 1]),
                "imag_values" : list(impedance_data[:, 2]),
                "nodal_attribution" : False,
                "averaged" : False  }

    # data_Z = {  "anechoic_termination": True,
    #             "volume_id": 1,
    #             "nodal_attribution": False  }

    model.set_surface_velocity(data_Vn, 1)
    model.set_specific_impedance(data_Z, 1)
    model.set_specific_impedance(data_Z, 2)

    # Define the analysis frequency setup
    df = 5
    f_min = 5
    f_max = 1400
    frequencies = np.arange(f_min, f_max + df, df)

    # Configure porous material
    # pm_data = get_porous_material_data(model=pm_model)
    # model.set_porous_material_model_data(pm_data, volume=1)
    # model.process_porous_material_properties(frequencies)

    assembler = AcousticAssembler(model)

    # Set the analysis frequency setup
    assembler.process_assemble()
    
    # t0 = time()
    # # Run modal analysis
    # modal_solver = AcousticModalSolver(assembler)
    # natural_frequencies, modal_shape = modal_solver.solve()
    # dt = time() - t0
    # print(f"Elapsed time to solve modal analysis: {round(dt, 4)}s")
    # return
    
    # Define the analysis type and load setup
    analysis_data = {"analysis_id" : 3, "frequencies" : frequencies}
    harmonic_solver = AcousticHarmonicSolver(assembler, analysis_data=analysis_data)

    # Run harmonic analysis

    t0 = time()
    solution = harmonic_solver.solve(print_log=True)
    dt = time() - t0
    print(f"Elapsed time to solve harmonic analysis: {round(dt, 4)}")

    t0 = time()

    element_3d, _ = assembler.get_element()
    element_3d.reorder_connect()

    list_nodes = list()
    for tag, surface_nodes in mesh.nodes_from_surfaces.items():
        list_nodes.extend(surface_nodes)

    rho_eff_v1 = model.get_fluid_density_for_particle_velocity_calculation(1, frequencies)
    rho_eff_v2 = model.get_fluid_density_for_particle_velocity_calculation(2, frequencies)

    input_particle_velocity = harmonic_solver.get_particle_velocity_from_surface(1, rho_eff_v1)
    output_particle_velocity = harmonic_solver.get_particle_velocity_from_surface(2, rho_eff_v2)

    input_velocities = np.array(list(input_particle_velocity["Vx"].values()), dtype=complex)
    output_velocities = np.array(list(output_particle_velocity["Vx"].values()), dtype=complex)

    input_Vx = np.average(input_velocities, axis=0)
    output_Vx = np.average(output_velocities, axis=0)

    solid_elements_connected_to_nodes =  mesh.get_solid_elements_connected_to_nodes(list_nodes)

    particle_velocity = dict()
    for _node_id, element_ids in solid_elements_connected_to_nodes.items():
        Vk = 0.
        for _element_id in element_ids:
            Vk += element_3d.process_particle_velocity(_element_id, _node_id, rho_eff_v1, frequencies, solution)
        particle_velocity[_node_id] = Vk / len(element_ids)

    # input_Vx = 0.
    # for node_id in mesh.nodes_from_surfaces[1]:
    #     input_Vx += particle_velocity[node_id][0, :]
    # input_Vx /= len(mesh.nodes_from_surfaces[1])

    # output_Vx = 0.
    # for node_id in mesh.nodes_from_surfaces[2]:
    #     output_Vx += particle_velocity[node_id][0, :]
    # output_Vx /= len(mesh.nodes_from_surfaces[2])

    mesh._process_face_elements_connected_to_nodes([1, 2])
    mesh._process_nodal_areas()

    freq_TL, TL_model, diff_TL = harmonic_solver.get_transmission_loss(1, 2)

    dt = time() - t0
    print(f"Elapsed time to post-process data: {round(dt, 4)}")

    if solution is not None:

        imported_results = get_external_results()
        
        pressure_at_input_face = imported_results["input_face_pressure"]
        pressure_at_output_face = imported_results["output_face_pressure"]
        velocity_at_input_face = imported_results["input_face_velocity"]
        velocity_at_output_face = imported_results["output_face_velocity"]
        pressure_at_node_6463 = imported_results["pressure_at_node_6463"]
        pressure_at_node_6531 = imported_results["pressure_at_node_6531"]
        velocity_at_node_6463 = imported_results["velocity_at_node_6463"]
        velocity_at_node_6531 = imported_results["velocity_at_node_6531"]
        TL_data = imported_results["transmission_loss"] # ports enabled

        output_ns = "output_face"

        if output_ns == "input_face":
            rows = mesh.nodes_from_surfaces[1]
            freq_ref = pressure_at_input_face[:, 0]
            results_ref = pressure_at_input_face[:, 1] + 1j*pressure_at_input_face[:, 2]

        else:
            rows = mesh.nodes_from_surfaces[2]
            freq_ref = pressure_at_output_face[:, 0]
            results_ref = pressure_at_output_face[:, 1] + 1j*pressure_at_output_face[:, 2]

        nodal_solution = np.average(solution[rows, :], axis=0).flatten()

        title = f"Harmonic response at {output_ns}"

        # abs_diff = np.max(np.abs((nodal_solution-results_ref)/results_ref))
        # print(f"Deviation: {100*abs_diff}")
        # assert abs_diff < 1e-4

        fig1, ax1 = plt.subplots()
        ax1.semilogy(frequencies, np.abs(nodal_solution), 'r', label='VIBRA')
        ax1.semilogy(freq_ref, np.abs(results_ref), 'k--', label='ANSYS')
        ax1.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Absolute', title=title)
        ax1.grid()
        ax1.legend()

        fig2, ax2 = plt.subplots()
        ax2.plot(frequencies, np.real(nodal_solution), 'r', label='VIBRA')
        ax2.plot(freq_ref, np.real(results_ref), 'k--', label='ANSYS')
        ax2.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title=title)
        ax2.grid()
        ax2.legend()

        fig3, ax3 = plt.subplots()
        ax3.plot(frequencies, np.imag(nodal_solution), 'r', label='VIBRA')
        ax3.plot(freq_ref, np.imag(results_ref), 'k--', label='ANSYS')
        ax3.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title=title)
        ax3.grid()
        ax3.legend()

        # Plot the nodal results for pressure and particle velocity

        data_type = np.real
        type_label = "real"

        x_data_WB = pressure_at_node_6463[:, 0]
        y_data_WB = pressure_at_node_6463[:, 1] + 1j*pressure_at_node_6463[:, 2]

        fig4, ax4 = plt.subplots()
        title = "Acoustic pressure at node 6463"
        ax4.plot(frequencies, data_type(solution[6463-1, :]), 'r', label='VIBRA')
        ax4.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax4.set_xlabel('Frequency [Hz]')
        ax4.set_ylabel(f'Acoustic Pressure [Pa] - {type_label}')
        ax4.set_title(title)
        ax4.grid()
        ax4.legend()

        x_data_WB = pressure_at_node_6531[:, 0]
        y_data_WB = pressure_at_node_6531[:, 1] + 1j*pressure_at_node_6531[:, 2]

        fig5, ax5 = plt.subplots()
        title = "Acoustic pressure at node 6531"
        ax5.plot(frequencies, data_type(solution[6531-1, :]), 'r', label='VIBRA')
        ax5.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax5.set_xlabel('Frequency [Hz]')
        ax5.set_ylabel(f'Acoustic Pressure [Pa] - {type_label}')
        ax5.set_title(title)
        ax5.grid()
        ax5.legend()

        x_data_WB = velocity_at_node_6463[:, 0]
        y_data_WB = velocity_at_node_6463[:, 1] + 1j*velocity_at_node_6463[:, 2]

        fig6, ax6 = plt.subplots()
        title = "Particle velocity at node 6463"
        ax6.plot(frequencies, data_type(particle_velocity[6463-1][0, :]), 'r', label='VIBRA')
        ax6.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax6.set_xlabel('Frequency [Hz]')
        ax6.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax6.set_title(title)
        ax6.grid()
        ax6.legend()

        x_data_WB = velocity_at_node_6531[:, 0]
        y_data_WB = velocity_at_node_6531[:, 1] + 1j*velocity_at_node_6531[:, 2]

        fig7, ax7 = plt.subplots()
        title = "Particle velocity at node 6531"
        ax7.plot(frequencies, data_type(particle_velocity[6531-1][0, :]), 'r', label='VIBRA')
        ax7.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax7.set_xlabel('Frequency [Hz]')
        ax7.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax7.set_title(title)
        ax7.grid()
        ax7.legend()

        x_data_WB = velocity_at_input_face[:, 0]
        y_data_WB = velocity_at_input_face[:, 1] + 1j*velocity_at_input_face[:, 2]

        fig8, ax8 = plt.subplots()
        title = "Input face particle velocity - average"
        ax8.plot(frequencies, data_type(input_Vx), 'r', label='VIBRA')
        ax8.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax8.set_xlabel('Frequency [Hz]')
        ax8.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax8.set_title(title)
        ax8.grid()
        ax8.legend()

        x_data_WB = velocity_at_output_face[:, 0]
        y_data_WB = velocity_at_output_face[:, 1] + 1j*velocity_at_output_face[:, 2]

        fig9, ax9 = plt.subplots()
        title = "Output face particle velocity - average"
        ax9.plot(frequencies, data_type(output_Vx), 'r', label='VIBRA')
        ax9.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax9.set_xlabel('Frequency [Hz]')
        ax9.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax9.set_title(title)
        ax9.grid()
        ax9.legend()

        # Sound intensity at input face node

        x_data_WB = velocity_at_node_6531[:, 0]
        Vx_6531_WB = velocity_at_node_6531[:, 1] + 1j*velocity_at_node_6531[:, 2]
        P_6531_WB = pressure_at_node_6531[:, 1] + 1j*pressure_at_node_6531[:, 2]

        sound_int = np.real(solution[6531-1, :] * np.conj(particle_velocity[6531-1][0, :])) / 2
        y_data_WB = np.real(P_6531_WB * np.conj(Vx_6531_WB)) / 2

        fig10, ax10 = plt.subplots()
        title = "Sound intensity at node 6531"
        ax10.plot(frequencies, sound_int, 'r', label='VIBRA')
        ax10.plot(x_data_WB, y_data_WB, 'k--', label='ANSYS')
        ax10.set_xlabel('Frequency [Hz]')
        ax10.set_ylabel(f'Sound intensity [Pa.m/s] - {type_label}')
        ax10.set_title(title)
        ax10.grid()
        ax10.legend()

        # Sound intensity at output face node

        x_data_WB = velocity_at_node_6463[:, 0]
        Vx_6463_WB = velocity_at_node_6463[:, 1] + 1j*velocity_at_node_6463[:, 2]
        P_6463_WB = pressure_at_node_6463[:, 1] + 1j*pressure_at_node_6463[:, 2]

        sound_int = np.real(solution[6463-1, :] * np.conj(particle_velocity[6463-1][0, :])) / 2
        y_data_WB = np.real(P_6463_WB * np.conj(Vx_6463_WB)) / 2

        fig11, ax11 = plt.subplots()
        title = "Sound intensity at node 6463"
        ax11.plot(frequencies, sound_int, 'r', label='VIBRA')
        ax11.plot(x_data_WB, y_data_WB, 'k--', label='ANSYS')
        ax11.set_xlabel('Frequency [Hz]')
        ax11.set_ylabel(f'Sound intensity [Pa.m/s] - {type_label}')
        ax11.set_title(title)
        ax11.grid()
        ax11.legend()

        fig12, ax12 = plt.subplots()
        title = "Transmission loss"
        x_data_WB = TL_data[:, 0]
        y_data_WB = TL_data[:, 1]
        ax12.plot(freq_TL, TL_model, 'r', label='VIBRA')
        ax12.plot(x_data_WB, data_type(y_data_WB), 'k--', label='ANSYS')
        ax12.set_xlabel('Frequency [Hz]')
        ax12.set_ylabel(f'Transmission loss [dB] - {type_label}')
        ax12.set_title(title)
        ax12.grid()
        ax10.legend()

        plt.show()


def get_porous_material_data(model="DB"):

    if model == "DB":

        material_model_data = {
                                "model" : "Delany-Bazley",
                                "C1" : 0.0497,
                                "C2" : -0.754,
                                "C3" : 0.0758,
                                "C4" : -0.732,
                                "C5" : 0.169,
                                "C6" : -0.595,
                                "C7" : 0.0858,
                                "C8" : -0.700,
                                "flow_resistivity" : 1518.5066
                                }

    if model == "DBM":

        material_model_data = {
                                "model" : "Delany-Bazley-Miki",
                                "C1" : 0.070,
                                "C2" : -0.632,
                                "C3" : 0.1070,
                                "C4" : -0.632,
                                "C5" : 0.1600,
                                "C6" : -0.618,
                                "C7" : 0.1090,
                                "C8" : -0.618,
                                "flow_resistivity" : 1518.5066
                                }

    elif model == "JCA":

        material_model_data = {
                                "model" : "Jhonson-Champoux-Allard",
                                "porosity" : 0.9,
                                "tortuosity" : 1.0,
                                "viscous_characteristic_length" : 77e-6,
                                "thermal_characteristic_length" : 159e-6,
                                "flow_resistivity" : 1518.5066
                               }
    elif model == "JCAL":

        material_model_data = {
                                "model" : "Jhonson-Champoux-Allard-Lafarge",
                                "porosity" : 0.9,
                                "tortuosity" : 1.0,
                                "viscous_characteristic_length" : 77e-6,
                                "thermal_characteristic_length" : 159e-6,
                                "flow_resistivity" : 1518.5066
                               }

    return material_model_data


def get_external_results():

    imported_results = dict()
    # results_path = f"validation/data/particle_velocity/results/silencer/WB_results_silencer_only_fluid_{pm_model}_Vn1_Z1_Z2_complex.xlsx"
    # results_path = f"validation/data/particle_velocity/results/silencer/WB_results_silencer_only_fluid_{pm_model}_Vn1_Z1_Z2_real.xlsx"
    results_path = f"validation/data/particle_velocity/results/silencer/WB_results_silencer_only_fluid_Vn1_Z1_Z2_complex.xlsx"
    # results_path = f"validation/data/particle_velocity/results/silencer/WB_results_silencer_only_fluid_Vn1_Z1_Z2_real.xlsx"

    if not os.path.exists(results_path):
        return imported_results

    wb = load_workbook(results_path)

    skiprows = 0

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1,2]
                                    ).to_numpy()
        except:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1]
                                    ).to_numpy()

        imported_results[sheetname] = sheet_data

    return imported_results



def get_complex_impedance_data():

    imported_results = dict()
    results_path = f"validation/data/particle_velocity/results/silencer/complex_fluid_properties_DB_model.xlsx"

    if not os.path.exists(results_path):
        return imported_results

    wb = load_workbook(results_path)

    skiprows = 0

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1,2]
                                    ).to_numpy()
        except:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1]
                                    ).to_numpy()

        imported_results[sheetname] = sheet_data

    return imported_results


if __name__ == "__main__":
    test_load_external_mesh_and_solve(reorder_nodes=False)