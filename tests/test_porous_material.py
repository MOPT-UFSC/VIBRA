from vibra.engine.properties.fluid import Fluid
from vibra.engine.mesher.mesh import Mesh
from vibra.engine.mesher.element_type import *
from vibra.engine.model import Model
from vibra.engine.assemblers.acoustic_assembler import AcousticAssembler
from vibra.engine.solvers.acoustic_modal_solver import AcousticModalSolver
from vibra.engine.solvers.acoustic_harmonic_solver import AcousticHarmonicSolver
from vibra.external_mesh.external_mesh_data import ExternalMeshData

import os
import openpyxl

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

from time import time


def test_load_external_mesh_and_solve():
    # return

    # start decoding the Ansys script file (ds.dat file or input file)

    # mesh_path = "validation/data/porous_materials/mesh_files/fluid_suction_silencer_first_stage.dat"
    mesh_path = "validation/data/porous_materials/mesh_files/suction_silencer_first_stage.dat"

    if not os.path.exists(mesh_path):
        return

    # define the known 'Named selections' from model
    named_selecion_to_tag = { 
                                "input_face" : 1,
                                "output_face" : 2
                            }

    t0 = time()
    external_mesh = ExternalMeshData()
    external_mesh.reset()
    external_mesh.read_file(mesh_path)
    external_mesh.set_named_selections(list(named_selecion_to_tag.keys()))
    external_mesh.decode_mesh_data_from_file()
    
    dt = time() - t0
    print(f"\n\nElapsed time to decode the external mesh data: {round(dt, 4)} s")

    mesh = Mesh()
    mesh.import_external_nodal_coordinates(external_mesh.nodal_coordinates, index_zero=True)
    mesh.import_external_connectivity(external_mesh.connectivity_arrays, index_zero=True, etype_tag=4)
    mesh.element_type = TETRAHEDRON_4

    for named_selection, surf_data in external_mesh.elements_from_named_selection.items():
        tag = named_selecion_to_tag[named_selection]
        mesh.elements_from_surface[tag] = surf_data["element_indexes"]
        mesh.connectivity_from_surfaces[tag] = surf_data["connectivity"] - 1
        ns_nodes = external_mesh.nodes_from_named_selection[named_selection]
        mesh.nodes_from_surfaces[tag] = np.array(ns_nodes, dtype=int) - 1

        mesh.volume_from_surface[tag] = [3]

    mesh.surfaces_from_volumes[1] = [1, 2]

    # if reorder_nodes:
    #     mesh._process_nodes_reordering()
    #     map_nodes_indexes = mesh.reordering.map_nodes_indexes

    # Define the fluid properties

    temperature = 298.15
    pressure = 122525
    
    rho_0 = 2.634167
    c_0 = 225.307464
    mu = 8.156509e-06
    Cp = 1664.133942
    kt = 1.741553e-02
    gamma = 1.120295
    molar_mass = 51.951533

    fluid = Fluid(  name = "Silencer suction stg1",
                    identifier = 1,
                    color = (200, 200, 200),
                    pressure = pressure,
                    temperature = temperature,
                    fluid_density = rho_0,
                    speed_of_sound = c_0,
                    isentropic_exponent = gamma,
                    thermal_conductivity = kt,
                    specific_heat_Cp = Cp,
                    dynamic_viscosity = mu,
                    molar_mass = molar_mass  )

    # Set the defined fluid
    model = Model()
    model.mesh =  mesh
    model.generated_mesh = True
    model.set_fluid(fluid, volume=1)
    model.set_fluid(fluid, volume=2)
    model.set_fluid(fluid, volume=3)

    # Normal surface velocity data
    data_Vn = { "real_values" : [1],
                "imag_values" : [0],
                "nodal_attribution" : False,
                "averaged" : False }

    # Impedance data
    Zo = fluid.impedance
    data_Z = {  "real_values" : [Zo],
                "imag_values" : [0],
                "nodal_attribution" : False,
                "averaged" : False  }

    model.set_surface_velocity(data_Vn, 1)
    model.set_specific_impedance(data_Z, 1)
    model.set_specific_impedance(data_Z, 2)

    # Define the analysis frequency setup
    df = 5
    f_min = 5
    f_max = 1400
    frequencies = np.arange(f_min, f_max + df, df)

    # Configure porous material
    pm_model = "DB"
    pm_data = get_porous_material_data(model=pm_model)
    # model.set_porous_material_model_data(pm_data, volume=1)
    model.set_porous_material_model_data(pm_data, volume=2)
    model.set_porous_material_model_data(pm_data, volume=3)
    model.process_porous_material_properties(frequencies)

    assembler = AcousticAssembler(model)

    # Set the analysis frequency setup
    assembler.set_frequencies(frequencies)

    assembler.process_assemble()
    
    # t0 = time()
    # # Run modal analysis
    # modal_solver = AcousticModalSolver(assembler)
    # natural_frequencies, modal_shape = modal_solver.solve()
    # dt = time() - t0
    # print(f"Elapsed time to solve modal analysis: {round(dt, 4)}s")
    # return
    
    # Define the analysis type and load setup
    analysis_data = {"analysis_id" : 3, "frequencies" : frequencies}
    harmonic_solver = AcousticHarmonicSolver(assembler, analysis_data=analysis_data)

    # Run harmonic analysis

    t0 = time()
    solution = harmonic_solver.solve(print_log=True)
    dt = time() - t0
    print(f"Elapsed time to solve harmonic analysis: {round(dt, 4)}")

    mesh._process_face_elements_connected_to_nodes([1, 2])
    mesh._process_nodal_areas()

    freq_TL, TL_model, diff_TL = harmonic_solver.get_transmission_loss(1, 2)

    if solution is not None:

        output_ns = "output_face"
        imported_results = import_results()

        if output_ns == "input_face":

            rows = mesh.nodes_from_surfaces[1]

            if pm_model == "DB":
                # data = imported_results["input_ns_DB"]
                # data = imported_results["input_ns_Z1_DB"]
                # data = imported_results["input_ns_Z2_DB"]
                # data = imported_results["input_pressure_DB_Vn_Z1"]
                data = imported_results["input_pressure_no_gap_DB"]
                TL_data = imported_results["TL_no_gap_DB"]

            elif pm_model == "DBM":
                data = imported_results["input_ns_DBM"]

            elif pm_model == "JCA":
                data = imported_results["input_ns_JCA"]

        elif output_ns == "output_face":

            rows = mesh.nodes_from_surfaces[2]

            if pm_model == "DB":
                # data = imported_results["output_ns_DB"]
                # data = imported_results["output_ns_Z1_DB"]
                # data = imported_results["output_ns_Z2_DB"]
                # data = imported_results["output_pressure_DB_Vn_Z1"]
                data = imported_results["output_pressure_no_gap_DB"]
                TL_data = imported_results["TL_no_gap_DB"]

            elif pm_model == "DBM":
                data = imported_results["output_ns_DBM"]

            else:
                data = imported_results["output_ns_JCA"]

        else:
            return

        nodal_solution = np.average(solution[rows, :], axis=0).flatten()

        freq_ref = data[:, 0]
        results_ref = data[:, 1] + 1j*data[:, 2]

        title = f"Harmonic response at {output_ns}"

        fig1, ax1 = plt.subplots()
        ax1.semilogy(frequencies, np.abs(nodal_solution), 'r', label='VIBRA')
        ax1.semilogy(freq_ref, np.abs(results_ref), 'k--', label='ANSYS')
        ax1.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Absolute', title=title)
        ax1.grid()
        ax1.legend()

        fig2, ax2 = plt.subplots()
        ax2.plot(frequencies, np.real(nodal_solution), 'r', label='VIBRA')
        ax2.plot(freq_ref, np.real(results_ref), 'k--', label='ANSYS')
        ax2.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title=title)
        ax2.grid()
        ax2.legend()

        fig3, ax3 = plt.subplots()
        ax3.plot(frequencies, np.imag(nodal_solution), 'r', label='VIBRA')
        ax3.plot(freq_ref, np.imag(results_ref), 'k--', label='ANSYS')
        ax3.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title=title)
        ax3.grid()
        ax3.legend()

        fig4, ax4 = plt.subplots()
        freq_ref = TL_data[:, 0]
        TL_ref = TL_data[:, 1]
        ax4.plot(freq_TL, TL_model, 'r', label='VIBRA')
        ax4.plot(freq_ref, TL_ref, 'k--', label='ANSYS')
        ax4.set(xlabel='Frequency [Hz]', ylabel='Transmission loss [dB]', title="Transmission loss")
        ax4.grid()
        ax4.legend()

        # fig5, ax5 = plt.subplots()
        # freq_ref = TL_data[:, 0]
        # TL_ref = TL_data[:, 1]
        # ax5.plot(freq_TL, diff_TL, 'r', label='VIBRA')
        # # ax5.plot(freq_ref, TL_ref, 'k--', label='ANSYS')
        # ax5.set(xlabel='Frequency [Hz]', ylabel='Transmission loss [dB]', title="Transmission loss")
        # ax5.grid()
        # ax5.legend()

        plt.show()


def get_porous_material_data(model="DB"):

    if model == "DB":

        material_model_data = {
                                "model" : "Delany-Bazley",
                                "C1" : 0.0497,
                                "C2" : -0.754,
                                "C3" : 0.0758,
                                "C4" : -0.732,
                                "C5" : 0.169,
                                "C6" : -0.595,
                                "C7" : 0.0858,
                                "C8" : -0.700,
                                "flow_resistivity" : 1518.5066
                                }

    if model == "DBM":

        material_model_data = {
                                "model" : "Delany-Bazley-Miki",
                                "C1" : 0.070,
                                "C2" : -0.632,
                                "C3" : 0.1070,
                                "C4" : -0.632,
                                "C5" : 0.1600,
                                "C6" : -0.618,
                                "C7" : 0.1090,
                                "C8" : -0.618,
                                "flow_resistivity" : 1518.5066
                                }

    elif model == "JCA":

        material_model_data = {
                                "model" : "Jhonson-Champoux-Allard",
                                "porosity" : 0.9,
                                "tortuosity" : 1.0,
                                "viscous_characteristic_length" : 77e-6,
                                "thermal_characteristic_length" : 159e-6,
                                "flow_resistivity" : 1518.5066
                               }
    elif model == "JCAL":

        material_model_data = {
                                "model" : "Jhonson-Champoux-Allard-Lafarge",
                                "porosity" : 0.9,
                                "tortuosity" : 1.0,
                                "viscous_characteristic_length" : 77e-6,
                                "thermal_characteristic_length" : 159e-6,
                                "flow_resistivity" : 1518.5066
                               }

    return material_model_data


def import_results():

    imported_results = dict()

    # results_path = "validation/data/porous_materials/results/porous_validation.xlsx"
    results_path = "validation/data/porous_materials/results/suction_silencer_1stg.xlsx"

    wb = openpyxl.load_workbook(results_path)

    skiprows = 0

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        try:
            sheet_data = pd.read_excel(results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1,2]).to_numpy()
        except:
            sheet_data = pd.read_excel(results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1]).to_numpy()

        imported_results[sheetname] = sheet_data

    return imported_results


def save_results(data):
    pass
    # node = 3596
    # if reorder_nodes:
    #     node = int(map_nodes_indexes[node])

    # cols = solution.shape[1]        
    # results = np.zeros((cols, 3), dtype=float)
    # results[:, 0] = frequencies
    # results[:, 1] = np.real(solution[node, :])
    # results[:, 2] = np.imag(solution[node, :])
    # filename = f"acoustic_pressure_at_node_{node}_Vibra_pardiso.dat"
    # np.savetxt(filename, results, delimiter=",")

    # results_path = "data/examples/mesh/muffler/external_results.csv"

    # data_ref = np.loadtxt(results_path, delimiter=",")
    # freq_ref = data_ref[:, 0]
    # P_ref = data_ref[:, 1] + 1j*data_ref[:, 2]

    # error_abs = np.abs((solution[node, :] - P_ref)/((solution[node, :] + P_ref)/2))
    # assert error_abs < 1e-1

# def get_faces_nodes():

#     input_face_path = "data/examples/mesh/porous_material/nodes_from_input_face.dat"
#     output_face_path = "data/examples/mesh/porous_material/nodes_from_output_face.dat"

#     input_face_data = np.loadtxt(input_face_path, delimiter=",", dtype=int) - 1
#     output_face_data = np.loadtxt(output_face_path, delimiter=",", dtype=int) - 1

#     nodes_from_surfaces = dict()

#     nodes_from_surfaces[1] = {  "node_indexes" : input_face_data[:, 0],
#                                        "nodes" : input_face_data[:, 1:]   }

#     nodes_from_surfaces[2] = {  "node_indexes" : output_face_data[:, 0],
#                                        "nodes" : output_face_data[:, 1:]   }

#     return nodes_from_surfaces


# def get_faces_connectivities():

#     input_face_path = "data/examples/mesh/porous_material/elements_from_input_face.dat"
#     output_face_path = "data/examples/mesh/porous_material/elements_from_output_face.dat"

#     input_face_data = np.loadtxt(input_face_path, delimiter=",", dtype=int) - 1
#     output_face_data = np.loadtxt(output_face_path, delimiter=",", dtype=int) - 1

#     connectivity_from_surfaces = dict()

#     connectivity_from_surfaces[1] = {   "element_indexes" : input_face_data[:, 0],
#                                         "connectivity" : input_face_data[:, 1:]   }

#     connectivity_from_surfaces[2] = {   "element_indexes" : output_face_data[:, 0],
#                                         "connectivity" : output_face_data[:, 1:]   }

#     return connectivity_from_surfaces

if __name__ == "__main__":
    test_load_external_mesh_and_solve(reorder_nodes=False)