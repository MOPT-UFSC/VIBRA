from vibra.engine.properties.material import Material
from vibra.engine.mesher.mesh import Mesh
from vibra.engine.mesher.element_type import *
from vibra.engine.model import Model

# from vibra.engine.assemblers.acoustic_assembler import AcousticAssembler
# from vibra.engine.solvers.acoustic_modal_solver import AcousticModalSolver
# from vibra.engine.solvers.acoustic_harmonic_solver import AcousticHarmonicSolver
from vibra.engine.assemblers.structural_assembler import StructuralAssembler
from vibra.engine.solvers.structural_modal_solver import StructuralModalSolver
from vibra.engine.solvers.structural_harmonic_solver import StructuralHarmonicSolver

from vibra.external_mesh.external_mesh_data import ExternalMeshData
from data.validation.load_external_data import LoadExternalData

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from vibra.engine.model import Model

import os
# import pytest
import numpy as np
import matplotlib.pyplot as plt

from time import time
from pandas import read_excel
from openpyxl import load_workbook

# valid mesh sizes: 20mm and 200mm.
mesh_size = "200mm"


# @pytest.mark.slow
# @pytest.mark.skip

def load_external_mesh_and_solve(interior_impedance: bool = False):

    # start decoding the Ansys script file (ds.dat file or input file)
    mesh_path = f"data/validation/structural/shell_element/pipes/ds_Lpipe_with_caps.dat"

    if not os.path.exists(mesh_path):
        return

    # define the known 'Named selections' from model
    named_selecion_to_tag = { 
                             "input_face" : 2,
                             "output_face" : 3,
                            }

    t0 = time()
    external_mesh = ExternalMeshData()
    external_mesh.reset()
    external_mesh.read_file(mesh_path)
    external_mesh.set_named_selections(list(named_selecion_to_tag.keys()))
    external_mesh.decode_mesh_data_from_file()

    # nodes_from_named_selection = external_mesh.nodes_from_named_selection
    # for ns, nodes in nodes_from_named_selection.items():
    #     print(ns, nodes)

    dt = time() - t0
    print(f"\n\nElapsed time to decode the external mesh data: {round(dt, 4)} s")

    mesh = Mesh()
    mesh.import_external_nodal_coordinates(external_mesh.nodal_coordinates, index_zero=True)
    mesh.import_external_faces_connectivity(external_mesh.connectivity_arrays, index_zero=True, etype_tag=4)
    mesh.export_nodal_coordinates("nodal_coordinates.dat")
    mesh.export_solid_elements_connectivity("solids_connectivity.dat")
    mesh.element_type = TETRAHEDRON_4

    for named_selection, surf_data in external_mesh.elements_from_named_selection.items():

        if named_selection in ["input_edges", "output_edges"]:
            continue

        tag = named_selecion_to_tag[named_selection]
        mesh.elements_from_surface[tag] = surf_data["element_indexes"] - 1
        mesh.connectivity_from_surfaces[tag] = surf_data["connectivity"] - 1
        mesh.nodes_out_of_face_element[tag] = surf_data["outer_nodes"] - 1
        ns_nodes = external_mesh.nodes_from_named_selection[named_selection]
        mesh.nodes_from_surfaces[tag] = np.array(ns_nodes, dtype=int) - 1


    # Define the material properties

    density = 7850
    elasticity_modulus = 2e11
    poisson_ratio = 0.30
    thermal_expansion_coefficient = 1.1e-5

    material = Material(   
                        name = "Carbon steel",
                        identifier = 1,
                        color = (200, 200, 200),
                        density = density,
                        elasticity_modulus = elasticity_modulus,
                        poisson_ratio = poisson_ratio,
                        thermal_expansion_coefficient = thermal_expansion_coefficient
                        )


    # # Load the external data
    # if interior_impedance:
    #     path = f"data/validation/elementar/interior_impedance/results/mesh_size_{mesh_size}"
    # else:
    #     path = f"data/validation/elementar/results/mesh_size_{mesh_size}"

    # ext_data = LoadExternalData(path, rho_0)

    # assign the created fluid
    model = Model()
    model.mesh =  mesh
    model.generated_mesh = True

    data_thick = {
                  "surface_thickness": 0.008,
                  "thickness_offset": "middle",
                  }
    
    for _surf_id in [1]:
        model.properties._set_property("material", material, surface=_surf_id)
        model.properties._set_property("surface_thickness", data_thick, surface=_surf_id)

    # Prescribed dofs data
    prescribed_dofs_data = {
                            "element_type": "2d_element",
                            "real_values": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                            "imag_values": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                            }

    model.properties._set_property("prescribed_dofs", prescribed_dofs_data, surface=2)


    ## Create an object of the Structural Assembler class
    assembler = StructuralAssembler(model)  


    ## Define the analysis type and frequency setup

    # df = 5
    # f_min = 5
    # f_max = 1400
    # frequencies = np.arange(f_min, f_max + df, df)

    # frequency_setup = {
    #                    "f_min" : f_min,
    #                    "f_max" : f_max,
    #                    "f_step" : df,
    #                    "frequencies" : frequencies
    #                    }

    # # Set the analysis frequency setup
    # model.set_frequency_setup(frequency_setup)

    # model.set_structural_element(assembler.get_element())
    # analysis_data = {"analysis_id" : 3, "frequencies" : frequencies}
    # harmonic_solver = StructuralHarmonicSolver(assembler, analysis_data=analysis_data)


    # Define the analysis type and load setup
    model.set_structural_element(assembler.get_element())
    analysis_data = {"analysis_id" : 2, "modes" : 40, "sigma_factor" : 1e-2}
    modal_solver = StructuralModalSolver(assembler, analysis_data=analysis_data)


    # Run modal analysis

    assembler.process_assemble()

    t0 = time()
    solution = modal_solver.solve()
    dt = time() - t0
    print(f"Elapsed time to solve modal analysis: {round(dt, 4)}\n\n")

    print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    print(":: PLOTTING THE OBTAINED RESULTS FOR NATURAL FREQUENCIES ::")
    print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::::::\n")

    for k, freq in enumerate(modal_solver.natural_frequencies):
        print(f"Mode: {k+1} ==> Natural frequency: {freq : .4f} Hz")

    return

    input_rows = mesh.nodes_from_surfaces[1]
    output_rows = mesh.nodes_from_surfaces[2]

    input_pressure = np.average(solution[input_rows, :], axis=0).flatten()
    output_pressure = np.average(solution[output_rows, :], axis=0).flatten()

    t0 = time()

    element_3d, _ = assembler.get_element()
    element_3d.reorder_connect()

    list_nodes = list()
    for tag, surface_nodes in mesh.nodes_from_surfaces.items():
        list_nodes.extend(surface_nodes)

    rho_eff_v1 = model.get_fluid_density_for_particle_velocity_calculation(1, frequencies)
    rho_eff_v2 = model.get_fluid_density_for_particle_velocity_calculation(2, frequencies)

    input_particle_velocity = harmonic_solver.get_particle_velocity_from_surface(1, rho_eff_v1)
    output_particle_velocity = harmonic_solver.get_particle_velocity_from_surface(2, rho_eff_v2)

    input_velocities = np.array(list(input_particle_velocity["Vx"].values()), dtype=complex)
    output_velocities = np.array(list(output_particle_velocity["Vx"].values()), dtype=complex)

    input_Vx = np.average(input_velocities, axis=0)
    output_Vx = np.average(output_velocities, axis=0)

    solid_elements_connected_to_nodes =  mesh.get_solid_elements_connected_to_nodes(list_nodes)

    particle_velocity = dict()
    for _node_id, element_ids in solid_elements_connected_to_nodes.items():
        Vk = 0.
        for _element_id in element_ids:
            Vk += element_3d.process_particle_velocity(_element_id, _node_id, rho_0, frequencies, solution)
        particle_velocity[_node_id] = Vk / len(element_ids)

    # # nodal area calculation
    # mesh._process_face_elements_connected_to_nodes([1, 2])
    # mesh._process_nodal_areas()

    # freq_TL, TL_model = harmonic_solver.get_transmission_loss(1, 2)

    dt = time() - t0
    print(f"Elapsed time to post-process data: {round(dt, 4)}")

    if solution is not None:

        if mesh_size == "200mm":
            node_in = 5
            node_out = 1

        elif mesh_size == "20mm":
            node_in = 42
            node_out = 73

        else:
            return

        # # import the WB transmission loss data from spreadsheet file
        # imported_results = get_external_results(interior_impedance)
        # TL_data = imported_results["transmission_loss"] # ports enabled

        WB_pressure_data = ext_data.load_nodal_pressures()
        WB_particle_velocities_data = ext_data.load_particle_velocities()
        # WB_nodal_area_data = load_nodal_area()

        freq_WB, _, input_velocities_WB = WB_particle_velocities_data["Vx", "input_face"]
        input_Vx_WB = np.average(list(input_velocities_WB.values()), axis=0)

        freq_WB, _, input_pressures_WB = WB_pressure_data["input_face"]
        input_pressure_WB = np.average(list(input_pressures_WB.values()), axis=0)

        freq_WB, _, output_velocities_WB = WB_particle_velocities_data["Vx", "output_face"]
        output_Vx_WB = np.average(list(output_velocities_WB.values()), axis=0)

        freq_WB, _, output_pressures_WB = WB_pressure_data["output_face"]
        output_pressure_WB = np.average(list(output_pressures_WB.values()), axis=0)

        # Print the nodal results deviations
        abs_diff_node_Pin = np.abs((input_pressures_WB[node_in] - solution[node_in-1, :]) / (input_pressures_WB[node_in]))
        print(f"\nDeviation of pressure (node {node_in}): {100 * np.max(abs_diff_node_Pin)} %")

        abs_diff_node_Pout = np.abs((output_pressures_WB[node_out] - solution[node_out-1, :]) / (output_pressures_WB[node_out]))
        print(f"Deviation of pressure (node {node_out}): {100 * np.max(abs_diff_node_Pout)} %")

        abs_diff_node_Vin = np.abs((input_velocities_WB[node_in] - particle_velocity[node_in-1][0, :]) / (input_velocities_WB[node_in]))
        print(f"Deviation of particle velocity (node {node_in}): {100 * np.max(abs_diff_node_Vin)} %")

        abs_diff_node_Vout = np.abs((output_velocities_WB[node_out] - particle_velocity[node_out-1][0, :]) / (output_velocities_WB[node_out]))
        print(f"Deviation of particle velocity (node {node_out}): {100 * np.max(abs_diff_node_Vout)} %")

        abs_diff_Pinput_face = np.abs((input_pressure_WB - input_pressure) / input_pressure_WB)
        print(f"Deviation of pressure (input face): {100 * np.max(abs_diff_Pinput_face)} %")

        abs_diff_Poutput_face = np.abs((output_pressure_WB - output_pressure) / output_pressure_WB)
        print(f"Deviation of pressure (output face): {100 * np.max(abs_diff_Poutput_face)} %")

        abs_diff_Vinput_face = np.abs((input_Vx_WB - input_Vx) / input_Vx_WB)
        print(f"Deviation of particle velocity (input face): {100 * np.max(abs_diff_Vinput_face)} %")

        abs_diff_Voutput_face = np.abs((output_Vx_WB - output_Vx) / output_Vx_WB)
        print(f"Deviation of particle velocity (output face): {100 * np.max(abs_diff_Voutput_face)} %")


        title = f"Harmonic response at input face"

        fig1, ax1 = plt.subplots()
        ax1.plot(frequencies, np.real(input_pressure), 'r', label='VIBRA')
        ax1.plot(freq_WB, np.real(input_pressure_WB), 'k--', label='ANSYS')
        ax1.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title=title)
        ax1.grid()
        ax1.legend()

        fig2, ax2 = plt.subplots()
        ax2.plot(frequencies, np.imag(input_pressure), 'r', label='VIBRA')
        ax2.plot(freq_WB, np.imag(input_pressure_WB), 'k--', label='ANSYS')
        ax2.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title=title)
        ax2.grid()
        ax2.legend()

        title = f"Harmonic response at output face"

        fig3, ax3 = plt.subplots()
        ax3.plot(frequencies, np.real(output_pressure), 'r', label='VIBRA')
        ax3.plot(freq_WB, np.real(output_pressure_WB), 'k--', label='ANSYS')
        ax3.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title=title)
        ax3.grid()
        ax3.legend()

        fig4, ax4 = plt.subplots()
        ax4.plot(frequencies, np.imag(output_pressure), 'r', label='VIBRA')
        ax4.plot(freq_WB, np.imag(output_pressure_WB), 'k--', label='ANSYS')
        ax4.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title=title)
        ax4.grid()
        ax4.legend()

        # Plot the nodal results for pressure and particle velocity

        data_type = np.real
        type_label = "imaginary"

        fig5, ax5 = plt.subplots()
        title = f"Acoustic pressure at node {node_in}"
        ax5.plot(frequencies, data_type(solution[node_in-1, :]), 'r', label='VIBRA')
        ax5.plot(freq_WB, data_type(input_pressures_WB[node_in]), 'k--', label='ANSYS')
        ax5.set_xlabel('Frequency [Hz]')
        ax5.set_ylabel(f'Acoustic Pressure [Pa] - {type_label}')
        ax5.set_title(title)
        ax5.grid()
        ax5.legend()

        fig6, ax6 = plt.subplots()
        title = f"Acoustic pressure at node {node_out}"
        ax6.plot(frequencies, data_type(solution[node_out-1, :]), 'r', label='VIBRA')
        ax6.plot(freq_WB, data_type(output_pressures_WB[node_out]), 'k--', label='ANSYS')
        ax6.set_xlabel('Frequency [Hz]')
        ax6.set_ylabel(f'Acoustic Pressure [Pa] - {type_label}')
        ax6.set_title(title)
        ax6.grid()
        ax6.legend()

        fig7, ax7 = plt.subplots()
        title = f"Particle velocity at node {node_in}"
        ax7.plot(frequencies, data_type(particle_velocity[node_in-1][0, :]), 'r', label='VIBRA')
        ax7.plot(freq_WB, data_type(input_velocities_WB[node_in]), 'k--', label='ANSYS')
        ax7.set_xlabel('Frequency [Hz]')
        ax7.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax7.set_title(title)
        ax7.grid()
        ax7.legend()

        fig8, ax8 = plt.subplots()
        title = f"Particle velocity at node {node_out}"
        ax8.plot(frequencies, data_type(particle_velocity[node_out-1][0, :]), 'r', label='VIBRA')
        ax8.plot(freq_WB, data_type(output_velocities_WB[node_out]), 'k--', label='ANSYS')
        ax8.set_xlabel('Frequency [Hz]')
        ax8.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax8.set_title(title)
        ax8.grid()
        ax8.legend()

        fig9, ax9 = plt.subplots()
        title = "Input face particle velocity - average"
        ax9.plot(frequencies, data_type(input_Vx), 'r', label='VIBRA')
        ax9.plot(freq_WB, data_type(input_Vx_WB), 'k--', label='ANSYS')
        ax9.set_xlabel('Frequency [Hz]')
        ax9.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax9.set_title(title)
        ax9.grid()
        ax9.legend()

        fig10, ax10 = plt.subplots()
        title = "Output face particle velocity - average"
        ax10.plot(frequencies, data_type(output_Vx), 'r', label='VIBRA')
        ax10.plot(freq_WB, data_type(output_Vx_WB), 'k--', label='ANSYS')
        ax10.set_xlabel('Frequency [Hz]')
        ax10.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax10.set_title(title)
        ax10.grid()
        ax10.legend()

        plt.show()


def get_external_results(interior_impedance: bool):

    imported_results = dict()
    if interior_impedance:
        results_path = f"data/validation/elementar/results/interior_impedance/connected_rectangular_cavities_{mesh_size}.xlsx"
    else:
        results_path = f"data/validation/elementar/results/connected_rectangular_cavities_{mesh_size}.xlsx"

    if not os.path.exists(results_path):
        return imported_results

    wb = load_workbook(results_path)

    skiprows = 0

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1,2]
                                    ).to_numpy()
        except:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1]
                                    ).to_numpy()

        imported_results[sheetname] = sheet_data

    return imported_results


if __name__ == "__main__":
    load_external_mesh_and_solve(interior_impedance=False)