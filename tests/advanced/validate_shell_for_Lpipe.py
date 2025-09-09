from vibra.engine.properties.material import Material
from vibra.engine.mesher.mesh import Mesh
from vibra.engine.mesher.element_type import TETRAHEDRON_4
from vibra.engine.model import Model

from vibra.engine.assemblers.structural_assembler import StructuralAssembler
from vibra.engine.solvers.structural_modal_solver import StructuralModalSolver
from vibra.engine.solvers.structural_harmonic_solver import StructuralHarmonicSolver

from vibra.external_mesh.external_mesh_data import ExternalMeshData
from data.validation.load_external_data import LoadExternalData

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from vibra.engine.model import Model

import os
# import pytest
import numpy as np
import matplotlib.pyplot as plt

from time import time
from pandas import read_excel
from openpyxl import load_workbook

# valid mesh sizes: 20mm and 200mm.
mesh_size = "200mm"


# @pytest.mark.slow
# @pytest.mark.skip

def load_external_mesh_and_solve():

    # start decoding the Ansys script file (ds.dat file or input file)
    mesh_path = f"data/validation/structural/shell/L_pipe/mesh/ds_Lpipe_with_caps.dat"

    if not os.path.exists(mesh_path):
        return

    # define the known 'Named selections' from model
    named_selecion_to_tag = { 
                             "input_face" : 2,
                             "output_face" : 3,
                            }

    t0 = time()
    external_mesh = ExternalMeshData()
    external_mesh.reset()
    external_mesh.read_file(mesh_path)
    external_mesh.set_named_selections(list(named_selecion_to_tag.keys()))
    external_mesh.decode_mesh_data_from_file()

    # nodes_from_named_selection = external_mesh.nodes_from_named_selection
    # for ns, nodes in nodes_from_named_selection.items():
    #     print(ns, nodes)

    # return

    dt = time() - t0
    print(f"\n\nElapsed time to decode the external mesh data: {round(dt, 4)} s")

    mesh = Mesh()
    mesh.import_external_nodal_coordinates(external_mesh.nodal_coordinates, index_zero=True)
    mesh.import_external_faces_connectivity(external_mesh.solids_connectivities, index_zero=True, etype_tag=4)
    mesh.export_nodal_coordinates("nodal_coordinates.dat")
    mesh.export_solid_elements_connectivity("solids_connectivity.dat")
    mesh.element_type = TETRAHEDRON_4

    for named_selection, surf_data in external_mesh.elements_from_named_selection.items():

        if named_selection in ["input_edges", "output_edges"]:
            continue

        tag = named_selecion_to_tag[named_selection]
        mesh.elements_from_surface[tag] = surf_data["element_indexes"] - 1
        mesh.external_connectivity_from_surfaces[tag] = surf_data["connectivity"] - 1
        mesh.nodes_out_of_face_element[tag] = surf_data["outer_nodes"] - 1
        ns_nodes = external_mesh.nodes_from_named_selection[named_selection]
        mesh.external_nodes_from_surfaces[tag] = np.array(ns_nodes, dtype=int) - 1


    # # Load the external data
    # path = f"data/validation/structural/shell/pipes/results/results_for_L_pipe.xlsx"
    # ext_data = LoadExternalData(path, rho_0)

    # assign the created fluid
    model = Model()
    model.mesh =  mesh
    model.generated_mesh = True

    data_thick = {
                  "surface_thickness": 0.008,
                  "thickness_offset": "middle",
                  }

    model.properties._set_property("surface_thickness", data_thick, surface=1)


    # Define the material properties

    density = 7850
    elasticity_modulus = 2e11
    poisson_ratio = 0.30
    thermal_expansion_coefficient = 1.1e-5

    material = Material(   
                        name = "Carbon steel",
                        identifier = 1,
                        color = (200, 200, 200),
                        material_density = density,
                        elasticity_modulus = elasticity_modulus,
                        poisson_ratio = poisson_ratio,
                        thermal_expansion_coefficient = thermal_expansion_coefficient
                        )

    model.properties._set_property("material", material, surface=1)


    # Prescribed dofs data
    prescribed_dofs_data = {
                            "element_type": "2d_element",
                            "real_values": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                            "imag_values": [0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                            }

    model.properties._set_property("prescribed_dofs", prescribed_dofs_data, surface=2)


    # Nodal loads data
    distributed_load_data = {
                             "element_type" : "2d_element",
                             "real_values" : [100.0, None, None],
                             "imag_values" : [0.0, None, None],
                             "unit" : "N/m²"
                             }

    model.properties._set_property("distributed_loads", distributed_load_data, surface=3)



    ## Define the analysis type and frequency setup

    df = 2
    f_min = 2
    f_max = 300
    frequencies = np.arange(f_min, f_max + df, df)

    analysis_setup = {  
                      "analysis_id" : 0,
                      "global_damping" : (1e-3, 1e-7, 0),
                      "f_min" : f_min,
                      "f_max" : f_max,
                      "f_step" : df,
                      "frequencies" : frequencies
                      }

    model.set_analysis_setup(analysis_setup)

    # harmonic_solver = StructuralHarmonicSolver(assembler)
    # # Define the analysis setup
    # analysis_setup = {
    #                   "analysis_id" : 2, 
    #                   "modes" : 40, 
    #                   "sigma_factor" : 1e-2
    #                   }
    
    # # Set the analysis setup
    # model.set_analysis_setup(analysis_setup)

    # Define and process the assemble
    assembler = StructuralAssembler(model)
    assembler.process_assemble()

    # Initialize the solver
    # modal_solver = StructuralModalSolver(assembler)
    harmonic_solver = StructuralHarmonicSolver(assembler)

    # t0 = time()
    # # solution = modal_solver.solve()
    # dt = time() - t0
    # print(f"Elapsed time to solve modal analysis: {round(dt, 4)}\n\n")

    # print("::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    # print(":: PLOTTING THE OBTAINED RESULTS FOR MODAL ANALYSIS ::")
    # print("::::::::::::::::::::::::::::::::::::::::::::::::::::::\n")

    # for k, freq in enumerate(modal_solver.natural_frequencies):
    #     print(f"Mode: {k+1} ==> Natural frequency: {freq : .4f} Hz")


    t0 = time()
    # solution = modal_solver.solve()
    harmonic_solver.solve_direct_method(print_log=True)
    dt = time() - t0
    print(f"Elapsed time to solve the analysis: {round(dt, 4)}")


    # print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
    # print(":: PLOTTING THE OBTAINED RESULTS FOR HARMONIC ANALYSIS ::")
    # print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::::\n")

    selected_nodes = mesh.external_nodes_from_surfaces[3]

    dofs_index = {
                  "ux" : 0,
                  "uy" : 1,
                  "uz" : 2,
                  "rx" : 3,
                  "ry" : 4,
                  "rz" : 5
                  }
    
    element_2d = model.structural_element_2d
    if element_2d is None:
        return

    dofs_per_node = element_2d.DOFS_PER_NODE
    gdofs = dofs_per_node * selected_nodes.reshape(-1, 1) + np.arange(dofs_per_node, dtype=int)

    ux_rows = gdofs[:, dofs_index["ux"]]
    uy_rows = gdofs[:, dofs_index["uy"]]
    solution = harmonic_solver.solution

    response_ux = np.average(solution[ux_rows, :], axis=0).flatten()
    response_uy = np.average(solution[uy_rows, :], axis=0).flatten()


    dt = time() - t0
    print(f"Elapsed time to post-process data: {round(dt, 4)}")

    if solution is not None:

        ## load external results data
        imported_results = get_external_results()
        output_face_ux_lin = imported_results[f"output_face_ux_lin"]
        output_face_ux_quad = imported_results[f"output_face_ux_quad"]

        output_face_uy_lin = imported_results[f"output_face_uy_lin"]
        output_face_uy_quad = imported_results[f"output_face_uy_quad"]

        freq_WB = output_face_ux_lin[:, 0]
        output_face_ux_lin_WB = output_face_ux_lin[:, 1] + 1j*output_face_ux_lin[:, 2]

        freq_WB = output_face_ux_quad[:, 0]
        output_face_ux_quad_WB = output_face_ux_quad[:, 1] + 1j*output_face_ux_quad[:, 2]

        freq_WB = output_face_uy_lin[:, 0]
        output_face_uy_lin_WB = output_face_uy_lin[:, 1] + 1j*output_face_uy_lin[:, 2]

        freq_WB = output_face_uy_quad[:, 0]
        output_face_uy_quad_WB = output_face_uy_quad[:, 1] + 1j*output_face_uy_quad[:, 2]


        title = f"Harmonic response at output face"

        fig1, ax1 = plt.subplots()
        ax1.semilogy(frequencies, np.abs(response_ux), 'r', label='VIBRA')
        ax1.semilogy(freq_WB, np.abs(output_face_ux_lin_WB), 'k--', label='ANSYS (lin.)')
        ax1.semilogy(freq_WB, np.abs(output_face_ux_quad_WB), 'b--', label='ANSYS (quad.)')
        ax1.set(xlabel='Frequency [Hz]', ylabel='Magnitude of displacement Ux [m]', title=title)
        ax1.grid()
        ax1.legend()

        fig2, ax2 = plt.subplots()
        ax2.semilogy(frequencies, np.abs(response_uy), 'r', label='VIBRA')
        ax2.semilogy(freq_WB, np.abs(output_face_uy_lin_WB), 'k--', label='ANSYS (lin.)')
        ax2.semilogy(freq_WB, np.abs(output_face_uy_quad_WB), 'b--', label='ANSYS (quad.)')
        ax2.set(xlabel='Frequency [Hz]', ylabel='Magnitude of displacement Uy [m]', title=title)
        ax2.grid()
        ax2.legend()

        plt.show()


def get_external_results():

    imported_results = dict()
    results_path = f"data/validation/structural/shell/L_pipe/results/results_for_L_pipe.xlsx"

    if not os.path.exists(results_path):
        return imported_results

    wb = load_workbook(results_path)

    skiprows = 0

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1,2]
                                    ).to_numpy()
        except:
            sheet_data = read_excel(
                                    results_path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1]
                                    ).to_numpy()

        imported_results[sheetname] = sheet_data

    return imported_results


if __name__ == "__main__":
    load_external_mesh_and_solve()