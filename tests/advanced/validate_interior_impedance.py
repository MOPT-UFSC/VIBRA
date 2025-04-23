from vibra.engine.properties.fluid import Fluid
from vibra.engine.mesher.mesh import Mesh
from vibra.engine.mesher.element_type import *
from vibra.engine.model import Model
from vibra.engine.assemblers.acoustic_assembler import AcousticAssembler
from vibra.engine.solvers.acoustic_modal_solver import AcousticModalSolver
from vibra.engine.solvers.acoustic_harmonic_solver import AcousticHarmonicSolver

from vibra.external_mesh.external_mesh_data import ExternalMeshData
from data.validation.load_external_data import LoadExternalData

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from vibra.engine.model import Model

import os
# import pytest
import numpy as np
import matplotlib.pyplot as plt

from time import time
from pandas import read_excel
from openpyxl import load_workbook

# valid mesh sizes: 10mm, 34mm, 200mm and 400mm.
mesh_size = "34mm"


# @pytest.mark.slow
# @pytest.mark.skip

def load_external_mesh_and_solve(interior_impedance: bool = False):

    # start decoding the Ansys script file (ds.dat file or input file)
    mesh_path = f"data/validation/perforated_plate/mesh/ds_connected_rectangular_cavities_{mesh_size}.dat"

    if not os.path.exists(mesh_path):
        return

    # define the known 'Named selections' from model
    named_selecion_to_tag = { 
                             "input_face" : 1,
                             "output_face" : 2,
                             "middle_face" : 3,
                             "input_connected_faces" : 4,
                             "output_connected_faces" : 5,
                            }

    # define surfaces from each volume
    surfaces_from_volume = { 1 : [1, 3, 4], 2 : [2, 5]}

    t0 = time()
    external_mesh = ExternalMeshData()
    external_mesh.reset()
    external_mesh.read_file(mesh_path)
    external_mesh.set_named_selections(list(named_selecion_to_tag.keys()))
    external_mesh.decode_mesh_data_from_file()

    # nodes_from_named_selection = external_mesh.nodes_from_named_selection
    # for ns, nodes in nodes_from_named_selection.items():
    #     print(ns, nodes)

    dt = time() - t0
    print(f"\n\nElapsed time to decode the external mesh data: {round(dt, 4)} s")

    mesh = Mesh()
    mesh.import_external_nodal_coordinates(external_mesh.nodal_coordinates, index_zero=True)
    mesh.import_external_solids_connectivity(external_mesh.connectivity_arrays, index_zero=True, etype_tag=4)
    mesh.export_nodal_coordinates("nodal_coordinates.dat")
    mesh.export_solid_elements_connectivity("solids_connectivity.dat")
    mesh.element_type = TETRAHEDRON_4

    for named_selection, surf_data in external_mesh.elements_from_named_selection.items():

        if named_selection in ["input_edges", "output_edges"]:
            continue

        tag = named_selecion_to_tag[named_selection]
        mesh.elements_from_surface[tag] = surf_data["element_indexes"] - 1
        mesh.connectivity_from_surfaces[tag] = surf_data["connectivity"] - 1
        mesh.nodes_out_of_face_element[tag] = surf_data["outer_nodes"] - 1
        ns_nodes = external_mesh.nodes_from_named_selection[named_selection]
        mesh.nodes_from_surfaces[tag] = np.array(ns_nodes, dtype=int) - 1

    for vol_id, surf_ids in surfaces_from_volume.items():
        for surf_id in surf_ids:
            mesh.volumes_from_surface[surf_id] = [vol_id]
        mesh.surfaces_from_volume[vol_id] = surf_ids

    # return
    # define the fluid properties
    temperature = 293.15
    pressure = 101325
    rho_0 = 1.204263
    c_0 = 343.395034
    mu = 1.8247e-5
    Cp = 1006.400178
    kt = 2.5503e-02
    gamma = 1.401985
    molar_mass = 28.958601

    fluid = Fluid(  name = "Air_20C",
                    identifier = 1,
                    color = (200, 200, 200),
                    pressure = pressure,
                    temperature = temperature,
                    fluid_density = rho_0,
                    speed_of_sound = c_0,
                    isentropic_exponent = gamma,
                    thermal_conductivity = kt,
                    specific_heat_Cp = Cp,
                    dynamic_viscosity = mu,
                    molar_mass = molar_mass  )
    
    # Load the external data
    if interior_impedance:
        path = f"data/validation/perforated_plate/results/interior_impedance/mesh_size_{mesh_size}"
    else:
        path = f"data/validation/perforated_plate/results/mesh_size_{mesh_size}"

    ext_data = LoadExternalData(path, rho_0)

    ## assign the created fluid
    model = Model()
    model.mesh =  mesh
    model.generated_mesh = True

    for _vol_id in [1, 2]:
        model.properties._set_property("fluid", fluid, volume=_vol_id)
    
    for _surf_id in [1, 2, 3]:
        model.properties._set_property("fluid", fluid, surface=_surf_id)

    ## normal surface velocity data
    data_Vn = { "real_values" : [1],
                "imag_values" : [0],
                "nodal_attribution" : False,
                "averaged" : False }

    model.properties._set_property("surface_velocity", data_Vn, surface=1)

    ## boundary impedance setup
    Zo = fluid.impedance
    data_Z = {  
              "real_values" : [Zo],
              "imag_values" : [0],
              "nodal_attribution" : False,
              "averaged" : False
              }

    model.properties._set_property("specific_impedance", data_Z, surface=1)
    model.properties._set_property("specific_impedance", data_Z, surface=2)

    ## interior impedance setup

    # if interior_impedance:
    #     data_Zin = {  
    #                 "real_values" : [2*10],
    #                 "imag_values" : [0],
    #                 "nodal_attribution" : False,
    #                 "averaged" : False
    #                 }

    #     model.properties._set_property("specific_impedance", data_Zin, surface=3)

    ## Define the analysis frequency setup

    df = 5
    f_min = 5
    f_max = 1400
    frequencies = np.arange(f_min, f_max + df, df)

    frequency_setup = {
                        "f_min" : f_min,
                        "f_max" : f_max,
                        "f_step" : df,
                        "frequencies" : frequencies
                       }
    
    model.set_frequency_setup(frequency_setup)

    ## Define the perforated plate setup

    if interior_impedance:
        pp_data = {
                   "formulation" : "circular_hole",
                   "plate_thickness" : 0.008,
                   "porosity" : 0.26,
                   "hole_diameter" : 0.005,
                   "discharge_coefficient" : 0.76,
                   }

        model.properties._set_property("perforated_plate_model", pp_data, surface=3)
        model.process_perforated_plate_impedance(frequencies)

    assembler = AcousticAssembler(model)

    # Set the analysis frequency setup
    assembler.process_assemble()

    # t0 = time()
    # # Run modal analysis
    # modal_solver = AcousticModalSolver(assembler)
    # modal_solver.solve()
    # natural_frequencies = modal_solver.natural_frequencies
    # modal_shape = modal_solver.solution
    # dt = time() - t0
    # print(f"Elapsed time to solve modal analysis: {round(dt, 4)}s")
    # return

    # Define the analysis type and load setup
    analysis_data = {"analysis_id" : 3, "frequencies" : frequencies}
    harmonic_solver = AcousticHarmonicSolver(assembler, analysis_data=analysis_data)

    # Run harmonic analysis

    t0 = time()
    solution = harmonic_solver.solve(print_log=True)
    dt = time() - t0
    print(f"Elapsed time to solve harmonic analysis: {round(dt, 4)}")

    input_rows = mesh.nodes_from_surfaces[1]
    output_rows = mesh.nodes_from_surfaces[2]

    input_pressure = np.average(solution[input_rows, :], axis=0).flatten()
    output_pressure = np.average(solution[output_rows, :], axis=0).flatten()

    t0 = time()

    element_3d, _ = assembler.get_element()
    element_3d.reorder_connect()

    list_nodes = list()
    for tag, surface_nodes in mesh.nodes_from_surfaces.items():
        list_nodes.extend(surface_nodes)

    rho_eff_v1 = model.get_fluid_density_for_particle_velocity_calculation(1, frequencies)
    rho_eff_v2 = model.get_fluid_density_for_particle_velocity_calculation(2, frequencies)

    input_particle_velocity = harmonic_solver.get_particle_velocity_from_surface(1, rho_eff_v1)
    output_particle_velocity = harmonic_solver.get_particle_velocity_from_surface(2, rho_eff_v2)

    input_velocities = np.array(list(input_particle_velocity["Vx"].values()), dtype=complex)
    output_velocities = np.array(list(output_particle_velocity["Vx"].values()), dtype=complex)

    input_Vx = np.average(input_velocities, axis=0)
    output_Vx = np.average(output_velocities, axis=0)

    solid_elements_connected_to_nodes =  mesh.get_solid_elements_connected_to_nodes(list_nodes)

    particle_velocity = dict()
    for _node_id, element_ids in solid_elements_connected_to_nodes.items():
        Vk = 0.
        for _element_id in element_ids:
            Vk += element_3d.process_particle_velocity(_element_id, _node_id, rho_0, frequencies, solution)
        particle_velocity[_node_id] = Vk / len(element_ids)

    # nodal area calculation
    mesh._process_face_elements_connected_to_nodes([1, 2])
    mesh._process_nodal_areas()

    freq_TL, TL_model = harmonic_solver.get_transmission_loss(1, 2)

    dt = time() - t0
    print(f"Elapsed time to post-process data: {round(dt, 4)}")

    if solution is not None:

        if mesh_size == "400mm":    
            node_in = 3
            node_out = 17

        elif mesh_size == "200mm":
            node_in = 28
            node_out = 34

        elif mesh_size == "34mm":
            node_in = 1179
            node_out = 327

        elif mesh_size == "10mm":
            node_in = 12802
            node_out = 1304

        else:
            return

        WB_pressure_data = ext_data.load_nodal_pressures()
        WB_particle_velocities_data = ext_data.load_particle_velocities()
        # WB_nodal_area_data = load_nodal_area()

        freq_WB, _, input_velocities_WB = WB_particle_velocities_data["Vx", "input_face"]
        input_Vx_WB = np.average(list(input_velocities_WB.values()), axis=0)

        freq_WB, _, input_pressures_WB = WB_pressure_data["input_face"]
        input_pressure_WB = np.average(list(input_pressures_WB.values()), axis=0)

        freq_WB, _, output_velocities_WB = WB_particle_velocities_data["Vx", "output_face"]
        output_Vx_WB = np.average(list(output_velocities_WB.values()), axis=0)

        freq_WB, _, output_pressures_WB = WB_pressure_data["output_face"]
        output_pressure_WB = np.average(list(output_pressures_WB.values()), axis=0)

        # Print the nodal results deviations
        abs_diff_node_Pin = np.abs((input_pressures_WB[node_in] - solution[node_in-1, :]) / (input_pressures_WB[node_in]))
        print(f"\nDeviation of pressure (node {node_in}): {100 * np.max(abs_diff_node_Pin)} %")

        abs_diff_node_Pout = np.abs((output_pressures_WB[node_out] - solution[node_out-1, :]) / (output_pressures_WB[node_out]))
        print(f"Deviation of pressure (node {node_out}): {100 * np.max(abs_diff_node_Pout)} %")

        abs_diff_node_Vin = np.abs((input_velocities_WB[node_in] - particle_velocity[node_in-1][0, :]) / (input_velocities_WB[node_in]))
        print(f"Deviation of particle velocity (node {node_in}): {100 * np.max(abs_diff_node_Vin)} %")

        abs_diff_node_Vout = np.abs((output_velocities_WB[node_out] - particle_velocity[node_out-1][0, :]) / (output_velocities_WB[node_out]))
        print(f"Deviation of particle velocity (node {node_out}): {100 * np.max(abs_diff_node_Vout)} %")

        abs_diff_Pinput_face = np.abs((input_pressure_WB - input_pressure) / input_pressure_WB)
        print(f"Deviation of pressure (input face): {100 * np.max(abs_diff_Pinput_face)} %")

        abs_diff_Poutput_face = np.abs((output_pressure_WB - output_pressure) / output_pressure_WB)
        print(f"Deviation of pressure (output face): {100 * np.max(abs_diff_Poutput_face)} %")

        abs_diff_Vinput_face = np.abs((input_Vx_WB - input_Vx) / input_Vx_WB)
        print(f"Deviation of particle velocity (input face): {100 * np.max(abs_diff_Vinput_face)} %")

        abs_diff_Voutput_face = np.abs((output_Vx_WB - output_Vx) / output_Vx_WB)
        print(f"Deviation of particle velocity (output face): {100 * np.max(abs_diff_Voutput_face)} %")


        title = f"Harmonic response at input face"

        fig1, ax1 = plt.subplots()
        ax1.plot(frequencies, np.real(input_pressure), 'r', label='Vibra')
        ax1.plot(freq_WB, np.real(input_pressure_WB), 'k--', label='Ansys')
        ax1.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title=title)
        ax1.grid()
        ax1.legend()

        fig2, ax2 = plt.subplots()
        ax2.plot(frequencies, np.imag(input_pressure), 'r', label='Vibra')
        ax2.plot(freq_WB, np.imag(input_pressure_WB), 'k--', label='Ansys')
        ax2.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title=title)
        ax2.grid()
        ax2.legend()

        title = f"Harmonic response at output face"

        fig3, ax3 = plt.subplots()
        ax3.plot(frequencies, np.real(output_pressure), 'r', label='Vibra')
        ax3.plot(freq_WB, np.real(output_pressure_WB), 'k--', label='Ansys')
        ax3.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Real', title=title)
        ax3.grid()
        ax3.legend()

        fig4, ax4 = plt.subplots()
        ax4.plot(frequencies, np.imag(output_pressure), 'r', label='Vibra')
        ax4.plot(freq_WB, np.imag(output_pressure_WB), 'k--', label='Ansys')
        ax4.set(xlabel='Frequency [Hz]', ylabel='Acoustic Pressure [Pa] - Imaginary', title=title)
        ax4.grid()
        ax4.legend()

        # Plot the nodal results for pressure and particle velocity

        data_type = np.real
        type_label = "imaginary"

        fig5, ax5 = plt.subplots()
        title = f"Acoustic pressure at node {node_in}"
        ax5.plot(frequencies, data_type(solution[node_in-1, :]), 'r', label='Vibra')
        ax5.plot(freq_WB, data_type(input_pressures_WB[node_in]), 'k--', label='Ansys')
        ax5.set_xlabel('Frequency [Hz]')
        ax5.set_ylabel(f'Acoustic Pressure [Pa] - {type_label}')
        ax5.set_title(title)
        ax5.grid()
        ax5.legend()

        fig6, ax6 = plt.subplots()
        title = f"Acoustic pressure at node {node_out}"
        ax6.plot(frequencies, data_type(solution[node_out-1, :]), 'r', label='Vibra')
        ax6.plot(freq_WB, data_type(output_pressures_WB[node_out]), 'k--', label='Ansys')
        ax6.set_xlabel('Frequency [Hz]')
        ax6.set_ylabel(f'Acoustic Pressure [Pa] - {type_label}')
        ax6.set_title(title)
        ax6.grid()
        ax5.legend()

        fig7, ax7 = plt.subplots()
        title = f"Particle velocity at node {node_in}"
        ax7.plot(frequencies, data_type(particle_velocity[node_in-1][0, :]), 'r', label='Vibra')
        ax7.plot(freq_WB, data_type(input_velocities_WB[node_in]), 'k--', label='Ansys')
        ax7.set_xlabel('Frequency [Hz]')
        ax7.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax7.set_title(title)
        ax7.grid()
        ax7.legend()

        fig8, ax8 = plt.subplots()
        title = f"Particle velocity at node {node_out}"
        ax8.plot(frequencies, data_type(particle_velocity[node_out-1][0, :]), 'r', label='Vibra')
        ax8.plot(freq_WB, data_type(output_velocities_WB[node_out]), 'k--', label='Ansys')
        ax8.set_xlabel('Frequency [Hz]')
        ax8.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax8.set_title(title)
        ax8.grid()
        ax8.legend()

        fig9, ax9 = plt.subplots()
        title = "Input face particle velocity - average"
        ax9.plot(frequencies, data_type(input_Vx), 'r', label='Vibra')
        ax9.plot(freq_WB, data_type(input_Vx_WB), 'k--', label='Ansys')
        ax9.set_xlabel('Frequency [Hz]')
        ax9.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax9.set_title(title)
        ax9.grid()
        ax9.legend()

        fig10, ax10 = plt.subplots()
        title = "Output face particle velocity - average"
        ax10.plot(frequencies, data_type(output_Vx), 'r', label='Vibra')
        ax10.plot(freq_WB, data_type(output_Vx_WB), 'k--', label='Ansys')
        ax10.set_xlabel('Frequency [Hz]')
        ax10.set_ylabel(f'Particle velocity [m/s] - {type_label}')
        ax10.set_title(title)
        ax10.grid()
        ax10.legend()

        # Transmission loss

        if interior_impedance:
            results_path = f"data/validation/perforated_plate/results/interior_impedance/connected_rectangular_cavities_transmission_loss.xlsx"
        else:
            results_path = f"data/validation/perforated_plate/results/connected_rectangular_cavities_transmission_loss.xlsx"

        # import the WB transmission loss data from spreadsheet file
        results_WB = get_external_results(results_path)
        TL_data_WB = results_WB[f"transmission_loss_{mesh_size}"] # ports enabled

        freq_WB_evaluated, TL_WB_evaluated = process_external_TL(model, ext_data)
        
        mask = TL_data_WB[:, 0] <= f_max
        freq_WB_direct = TL_data_WB[:, 0][mask]
        TL_WB_direct = TL_data_WB[:, 1][mask]

        freq_WB_direct = TL_data_WB[:, 0]
        TL_WB_direct = TL_data_WB[:, 1]

        fig11, ax11 = plt.subplots()
        title = "Transmission loss"
        ax11.plot(freq_TL, TL_model, 'r', label='Vibra')
        ax11.plot(freq_WB_direct, TL_WB_direct, 'k--', label='Ansys')
        ax11.plot(freq_WB_evaluated, TL_WB_evaluated, 'b--', label='Ansys (ext.)')

        # #TODO: remove this
        # path = f"data/validation/perforated_plate/results/TL_cavidades_retangulares_comsol_{mesh_size}.xlsx"
        # if os.path.exists(path):
        #     results_Comsol = get_external_results(path)
        #     TL_data_Comsol = results_Comsol["transmission_loss"]

        #     freq_Comsol = TL_data_Comsol[:, 0]
        #     TL_Comsol = TL_data_Comsol[:, 1]

        #     ax11.plot(freq_Comsol, TL_Comsol, 'g--', label='Comsol')

        ax11.set_xlabel('Frequency [Hz]')
        ax11.set_ylabel(f'Transmission loss [dB]')
        ax11.set_title(title)
        ax11.grid()
        ax11.legend()

        plt.show()

def process_external_TL(model: "Model", ext_data: LoadExternalData):

    input_surface_id = 1
    output_surface_id = 2

    A_in = model.mesh.surface_area_from_element_integration[input_surface_id]
    A_out = model.mesh.surface_area_from_element_integration[output_surface_id]

    WB_nodal_area_data = ext_data.load_nodal_area()
    WB_pressure_data = ext_data.load_nodal_pressures()
    WB_particle_velocity_data = ext_data.load_particle_velocities()

    _, nodal_area_input = WB_nodal_area_data["input_face"]
    _, nodal_area_output = WB_nodal_area_data["output_face"]

    freq_WB, _, pressures_input = WB_pressure_data["input_face"]
    freq_WB, _, pressures_output = WB_pressure_data["output_face"]

    freq_WB, _, particle_velocity_input = WB_particle_velocity_data["Vx", "input_face"]
    freq_WB, _, particle_velocity_output = WB_particle_velocity_data["Vx", "output_face"]

    keys_na_in = list(nodal_area_input.keys())
    keys_pr_in = list(pressures_input.keys())
    keys_pv_in = list(particle_velocity_input.keys())

    keys_na_out = list(nodal_area_output.keys())
    keys_pr_out = list(pressures_output.keys())
    keys_pv_out = list(particle_velocity_output.keys())

    if (keys_na_in == keys_pr_in == keys_pv_in) and (keys_na_out == keys_pr_out == keys_pv_out):

        surf_velocity = model.properties._get_property("surface_velocity", surface=input_surface_id)
        if isinstance(surf_velocity, dict):
            if "real_values" in surf_velocity.keys():
                real_values = np.array(surf_velocity["real_values"])
                imag_values = np.array(surf_velocity["imag_values"])
                V_in = real_values + 1j * imag_values
            else:
                return None, None

        specific_impedance = model.properties._get_property("specific_impedance", surface=input_surface_id)
        if isinstance(specific_impedance, dict):
            if "real_values" in specific_impedance.keys():
                real_values = np.array(specific_impedance["real_values"])
                imag_values = np.array(specific_impedance["imag_values"])
                Zo_in = real_values + 1j * imag_values

            elif "anechoic_termination" in specific_impedance.keys():

                pm_active, rho_eff_pm, C_eff_pm = model.is_porous_material_model_active(input_surface_id)
                tv_active, rho_eff_tv, C_eff_tv = model.is_viscous_thermal_model_active(input_surface_id)

                if pm_active:
                    density = rho_eff_pm
                    speed_of_sound = C_eff_pm

                elif tv_active:
                    density = rho_eff_tv
                    speed_of_sound = C_eff_tv

                else:
                    fluid = model.properties._get_property("fluid", surface=input_surface_id)
                    fluid: Fluid
                    density = fluid.fluid_density
                    speed_of_sound = fluid.speed_of_sound

                Zo_in = density * speed_of_sound

            else:
                Zo_in = specific_impedance["values"]

        else:
            return None, None

        P_downstream = V_in * Zo_in / 2
        V_downstream = P_downstream / Zo_in

        # P_in = np.array(list(pressures_input.values()), dtype=complex)
        # Vx_in = -np.array(list(particle_velocity_input.values()), dtype=complex)

        # P_downstream = (P_in + Zo_in * Vx_in) / 2
        # V_downstream = P_downstream / Zo_in

        I_in = np.real(P_downstream * np.conjugate(V_downstream)) / 2
        NA_in = np.array(list(nodal_area_input.values()), dtype=float).reshape(-1, 1)

        P_out = np.array(list(pressures_output.values()), dtype=complex)
        Vx_out = np.array(list(particle_velocity_output.values()), dtype=complex)

        I_out = np.real(P_out * np.conjugate(Vx_out)) / 2
        NA_out = np.array(list(nodal_area_output.values()), dtype=float).reshape(-1, 1)

        W_in = 10 * np.log10(np.sum(I_in * NA_in, axis=0) / 1e-12)
        W_out = 10 * np.log10(np.sum(I_out * NA_out, axis=0) / 1e-12)

        # print(f"Incident power: {W_in}[dB]")

        TL = W_in - W_out

        return freq_WB, TL


def get_external_results(path: str):

    imported_results = dict()

    if not os.path.exists(path):
        return imported_results

    wb = load_workbook(path)

    skiprows = 0

    sheetnames = wb.sheetnames
    for sheetname in sheetnames:

        try:
            sheet_data = read_excel(
                                    path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1,2]
                                    ).to_numpy()
        except:
            sheet_data = read_excel(
                                    path, 
                                    sheet_name = sheetname, 
                                    header = skiprows, 
                                    usecols = [0,1]
                                    ).to_numpy()

        imported_results[sheetname] = sheet_data

    return imported_results


if __name__ == "__main__":
    load_external_mesh_and_solve(interior_impedance=True)